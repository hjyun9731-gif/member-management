# V20 fresh package marker: 2026-08-31 cutover audit fix
from __future__ import annotations

import csv
import calendar
import hashlib
import io
import json
import re
import threading
import time
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Optional
from urllib.parse import quote
from zoneinfo import ZoneInfo

from fastapi import APIRouter, BackgroundTasks, Depends, File, Form, HTTPException, Query, UploadFile
from fastapi.responses import FileResponse, StreamingResponse
from pydantic import BaseModel, Field
from openpyxl import Workbook
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import and_, case, func, or_, inspect, text
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from app import models
from app.auth import get_current_user
from app.database import SessionLocal, get_db, engine
from app.excel_utils import is_association_member
from app.receivables_models import (
    ReceivableCharge,
    ReceivableContactLog,
    ReceivablePayment,
    ReceivableProfile,
    ReceivableSystemState,
    ReceivableImportBatch,
    ReceivableImportRow,
)

router = APIRouter()
STATIC_DIR = Path(__file__).resolve().parent.parent / "static"
DATA_FILE = Path(__file__).resolve().parent.parent / "data" / "legacy_receivables_2026.json"
KST = ZoneInfo("Asia/Seoul")

ACCOUNT_FEES = {"협회비": 10000, "관리비": 5000, "70세": 5000}
CONTACT_STATUSES = {"미연락", "연락완료", "부재", "재연락 필요", "문자발송"}

# 기존 MUSTARD/엑셀 원장은 2026년 8월까지의 "월말 미수금"을 이미 포함한다.
# 따라서 1~8월 자동부과를 다시 더하면 이중부과가 된다.
LEGACY_YEAR = 2026
LEGACY_DATA_THROUGH_MONTH = 8
LEGACY_DATA_THROUGH_KEY = "2026-08"
LEGACY_DATA_THROUGH_DATE = date(2026, 8, 31)
LEGACY_DATA_THROUGH_DATE_ISO = LEGACY_DATA_THROUGH_DATE.isoformat()
LEGACY_NEXT_BILL_DATE = date(2026, 9, 1)
LEGACY_NEW_MEMBER_CUTOFF = date(2026, 8, 1)
LEGACY_BASELINE_KEY = "legacy_baseline_2026_08_28_v1"
LEDGER_MODE = "database"

# 미수금(협회비/관리비/70세)과 무관한 수입은 자동수납 금지.
# 특히 자격증명 발급비를 관리비 선납으로 잡는 사고를 막기 위해 업로드/수동입력 양쪽에서 방어한다.
NON_RECEIVABLE_IMPORT_KEYWORDS = (
    "자격증명발급비", "자격증명 발급비", "자격증명발급", "자격증명 발급", "자격증명",
    "자격증발급비", "자격증 발급비", "발급수수료",
    "대폐차수수료", "대폐차 수수료", "대폐차비", "대폐차",
    "가입비",
)
NON_RECEIVABLE_REPAIR_KEY = "receivables_nonreceivable_import_repair_20260831_v2"
FEE_REPAIR_KEY = "receivables_fixed_monthly_fee_repair_20260831_v1"
# 2026년에는 기존 택배(차량번호에 "배") 관리비만 월 5,000원 자동부과한다.
# 비택배 일반 관리비는 2027-01-01부터 시작한다.
GENERAL_MANAGEMENT_START_DATE = date(2027, 1, 1)
GENERAL_MANAGEMENT_START_KEY = "2027-01"
MANAGEMENT_2027_REPAIR_KEY = "receivables_general_management_2027_start_repair_20260901_v1"

# receivables 전용 lazy schema guard.
# Railway healthcheck를 빠르게 통과시키기 위해 main.py의 전체 DB 유지보수는
# 백그라운드에서 돌지만, 사용자가 /receivables를 먼저 열어도 이 모듈의
# 수납/미수금 전용 테이블만 즉시 준비되도록 한다. 전체회원관리 테이블/로직은 건드리지 않는다.
_receivables_schema_ready = False
_receivables_schema_lock = threading.Lock()


def _ensure_receivables_schema_ready() -> None:
    global _receivables_schema_ready
    if _receivables_schema_ready:
        return
    with _receivables_schema_lock:
        if _receivables_schema_ready:
            return

        tables = [
            ReceivableProfile.__table__,
            ReceivableCharge.__table__,
            ReceivablePayment.__table__,
            ReceivableContactLog.__table__,
            ReceivableSystemState.__table__,
            ReceivableImportBatch.__table__,
            ReceivableImportRow.__table__,
        ]
        # 이 모듈 테이블만 checkfirst=True로 생성한다. 기존 회원/인허가 테이블은 대상 아님.
        ReceivableProfile.metadata.create_all(bind=engine, tables=tables, checkfirst=True)

        # create_all은 기존 테이블의 누락 컬럼을 ALTER하지 않으므로,
        # 과거 receivable_profiles가 이미 존재하는 경우 이 컬럼만 안전하게 보강한다.
        try:
            cols = {c['name'] for c in inspect(engine).get_columns('receivable_profiles')}
            if 'account_manual_override' not in cols:
                with engine.begin() as conn:
                    conn.execute(text(
                        'ALTER TABLE receivable_profiles '
                        'ADD COLUMN account_manual_override INTEGER DEFAULT 0'
                    ))
        except Exception:
            # main.py의 기존 migration도 동일 컬럼을 보강하므로 동시 실행 경쟁은 무시 가능.
            pass

        _receivables_schema_ready = True


class PaymentIn(BaseModel):
    payment_date: str
    amount: int = Field(gt=0)
    method: Optional[str] = None
    memo: Optional[str] = None


class ContactIn(BaseModel):
    contact_date: str
    contact_method: str = "전화"
    status: str = "연락완료"
    memo: Optional[str] = None


class AccountIn(BaseModel):
    account_type: str
    vehicle_count: int = Field(default=1, ge=1, le=1)


class BalanceEditIn(BaseModel):
    balance_type: str
    amount: int = Field(default=0, ge=0, le=1_000_000_000)
    effective_date: Optional[str] = None
    reason: str = Field(min_length=2, max_length=500)


def _norm(v) -> str:
    return re.sub(r"[^0-9A-Za-z가-힣]", "", str(v or "")).lower()


def _norm_vehicle_key(v) -> str:
    """차량번호 표기 차이(강원/호/공백/하이픈)를 제거한 비교키."""
    s = _norm(v)
    s = re.sub(r"^(강원특별자치도|강원도|강원)", "", s)
    s = re.sub(r"호$", "", s)
    return s


def _vehicle_tail4(v) -> str:
    digits = re.sub(r"\D", "", str(v or ""))
    return digits[-4:] if len(digits) >= 4 else ""


def _korean_only_name(v) -> str:
    return re.sub(r"[^가-힣]", "", str(v or ""))


def _company_name_key(v) -> str:
    s = _norm(v)
    for token in ("주식회사", "유한회사", "합자회사", "합명회사", "협동조합", "대표이사", "대표자", "대표", "법인", "주"):
        s = s.replace(token, "")
    return s


def _edit_distance_one(a: str, b: str) -> bool:
    if a == b:
        return True
    if abs(len(a) - len(b)) > 1:
        return False
    if len(a) == len(b):
        return sum(x != y for x, y in zip(a, b)) <= 1
    if len(a) > len(b):
        a, b = b, a
    i = j = diff = 0
    while i < len(a) and j < len(b):
        if a[i] == b[j]:
            i += 1; j += 1
        else:
            diff += 1; j += 1
            if diff > 1:
                return False
    return True


def _names_safely_equivalent(a, b) -> bool:
    """지역+차량번호가 이미 맞는 후보에서만 사용하는 보수적 성명 정규화."""
    na, nb = _norm(a), _norm(b)
    if not na or not nb:
        return False
    if na == nb:
        return True

    ka, kb = _korean_only_name(a), _korean_only_name(b)
    if ka and kb and ka == kb:
        return True

    ca, cb = _company_name_key(a), _company_name_key(b)
    if len(ca) >= 3 and len(cb) >= 3 and (ca.startswith(cb) or cb.startswith(ca)):
        return True

    # 외국인/영문 병기 중 실제 한글 이름이 포함된 경우.
    if ka and kb and len(ka) >= 2 and (ka in kb or kb in ka):
        return True

    aliases = {
        "ojimamieko": "오지마미에꼬",
    }
    if aliases.get(na) == kb or aliases.get(nb) == ka:
        return True

    # 오탈자 1글자는 지역+차량 끝4자리 후보가 유일할 때만 호출되므로 제한 허용.
    if ka and kb and len(ka) >= 3 and len(kb) >= 3 and _edit_distance_one(ka, kb):
        return True
    return False


def _parse_date(v) -> Optional[date]:
    """회원 DB의 2026-08-27 / 26.08.27 / 26-8-27 등을 모두 처리."""
    if not v:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v

    s = str(v).strip().rstrip(".")
    for fmt in (
        "%Y-%m-%d", "%Y.%m.%d", "%Y/%m/%d", "%Y%m%d",
        "%y-%m-%d", "%y.%m.%d", "%y/%m/%d",
        "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M:%S.%f",
    ):
        try:
            return datetime.strptime(s[:26], fmt).date()
        except Exception:
            pass

    m = re.search(r"(19\d{2}|20\d{2})\D*(\d{1,2})\D*(\d{1,2})", s)
    if m:
        try:
            return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except Exception:
            return None

    m = re.search(r"^(\d{2})\D+(\d{1,2})\D+(\d{1,2})", s)
    if m:
        yy = int(m.group(1))
        year = 2000 + yy if yy <= 30 else 1900 + yy
        try:
            return date(year, int(m.group(2)), int(m.group(3)))
        except Exception:
            return None
    return None


def _first_of_next_month(d: date) -> date:
    if d.month == 12:
        return date(d.year + 1, 1, 1)
    return date(d.year, d.month + 1, 1)


def _month_iter(start: date, end: date):
    cur = date(start.year, start.month, 1)
    end_m = date(end.year, end.month, 1)
    while cur <= end_m:
        yield cur
        cur = date(cur.year + (1 if cur.month == 12 else 0), 1 if cur.month == 12 else cur.month + 1, 1)


def _user_name(user) -> str:
    return getattr(user, "full_name", None) or getattr(user, "username", None) or "사용자"


def _is_active(member) -> bool:
    return (
        member is not None
        and getattr(member, "deleted_at", None) is None
        and (getattr(member, "status", None) or "active") == "active"
    )


def _active_sql():
    return and_(
        models.LicenseHolder.deleted_at.is_(None),
        or_(models.LicenseHolder.status.is_(None), models.LicenseHolder.status == "active"),
    )


def _closed_sql():
    return or_(
        models.LicenseHolder.deleted_at.isnot(None),
        and_(models.LicenseHolder.status.isnot(None), models.LicenseHolder.status != "active"),
    )


_seed_cache = None
_seed_by_combo = None
_seed_by_vehicle = None
_seed_by_name_region = None
_seed_by_name_plate_tail = None
_seed_by_region_plate_tail = None
_seed_by_region_vehicle_key = None
_seed_by_region_nonplate = None
_seed_by_source_row = None
_legacy_baseline_ready = False
_legacy_baseline_lock = threading.Lock()
_billing_ready_month = None
_billing_month_lock = threading.Lock()
_monthly_scheduler_started = False
_monthly_scheduler_lock = threading.Lock()


def _load_seed():
    global _seed_cache, _seed_by_combo, _seed_by_vehicle, _seed_by_name_region, _seed_by_name_plate_tail, _seed_by_region_plate_tail, _seed_by_region_vehicle_key, _seed_by_region_nonplate, _seed_by_source_row
    if _seed_cache is not None:
        return _seed_cache
    if not DATA_FILE.exists():
        _seed_cache = []
        _seed_by_combo, _seed_by_vehicle, _seed_by_name_region, _seed_by_name_plate_tail, _seed_by_region_plate_tail, _seed_by_region_vehicle_key, _seed_by_region_nonplate, _seed_by_source_row = {}, {}, {}, {}, {}, {}, {}, {}
        return _seed_cache

    payload = json.loads(DATA_FILE.read_text(encoding="utf-8"))
    rows = payload.get("rows", [])
    by_combo, by_vehicle, by_nr, by_name_tail, by_region_tail, by_region_vehicle, by_region_nonplate, by_source = {}, {}, {}, {}, {}, {}, {}, {}
    for r in rows:
        sr = r.get("source_row")
        if sr is not None:
            try:
                by_source[int(sr)] = r
            except Exception:
                pass
        nv, nn, nr = _norm(r.get("vehicle_number")), _norm(r.get("name")), _norm(r.get("region"))
        by_combo[(nv, nn)] = r
        if nv:
            by_vehicle.setdefault(nv, []).append(r)
        if nn:
            by_nr.setdefault((nn, nr), []).append(r)
            tail = _vehicle_tail4(r.get("vehicle_number"))
            if tail:
                by_name_tail.setdefault((nn, tail), []).append(r)
        tail = _vehicle_tail4(r.get("vehicle_number"))
        if nr and tail:
            by_region_tail.setdefault((nr, tail), []).append(r)
        vk = _norm_vehicle_key(r.get("vehicle_number"))
        if nr and vk:
            by_region_vehicle.setdefault((nr, vk), []).append(r)
        if nr and not tail:
            by_region_nonplate.setdefault(nr, []).append(r)
    _seed_cache, _seed_by_combo, _seed_by_vehicle, _seed_by_name_region, _seed_by_name_plate_tail, _seed_by_region_plate_tail, _seed_by_region_vehicle_key, _seed_by_region_nonplate, _seed_by_source_row = rows, by_combo, by_vehicle, by_nr, by_name_tail, by_region_tail, by_region_vehicle, by_region_nonplate, by_source
    return rows


def _match_seed(member):
    _load_seed()
    # Legacy 자동매칭 우선순위:
    # 1) 기존의 차량번호+성명 정확일치
    # 2) 성명+차량번호 끝4자리 유일
    # 3) 지역+차량번호(강원/호 제거) 후보 중 성명 표기만 달라진 유일 후보
    # 4) 지역+끝4자리 후보 중 성명 표기만 달라진 유일 후보
    # 5) 원장 차량번호 공란인 경우에만 성명+지역 유일
    #
    # 차량번호만 같고 성명이 전혀 다른 경우(양도/소유자 변경)는 절대 자동연결하지 않는다.
    nv, nn = _norm(member.vehicle_number), _norm(member.name)
    nr = _norm(getattr(member, "region", ""))
    if not nn:
        return None

    if nv:
        exact = _seed_by_combo.get((nv, nn))
        if exact:
            return exact

    tail = _vehicle_tail4(getattr(member, "vehicle_number", ""))
    if tail:
        candidates = _seed_by_name_plate_tail.get((nn, tail), [])
        if len(candidates) == 1:
            return candidates[0]

    # 사람 눈에는 같은 차량/이름인데 영문병기·대표자명·(주) 표기 때문에 기존 _norm이
    # 달라지던 케이스를 안전하게 복구한다. 반드시 지역+차량 후보를 먼저 좁힌다.
    vehicle_key = _norm_vehicle_key(getattr(member, "vehicle_number", ""))
    if nr and vehicle_key:
        candidates = _seed_by_region_vehicle_key.get((nr, vehicle_key), [])
        name_matches = [r for r in candidates if _names_safely_equivalent(member.name, r.get("name"))]
        if len(name_matches) == 1:
            return name_matches[0]

    if nr and tail:
        candidates = _seed_by_region_plate_tail.get((nr, tail), [])
        name_matches = [r for r in candidates if _names_safely_equivalent(member.name, r.get("name"))]
        if len(name_matches) == 1:
            return name_matches[0]

    if nr:
        nonplate_candidates = _seed_by_region_nonplate.get(nr, [])
        name_matches = [r for r in nonplate_candidates if _names_safely_equivalent(member.name, r.get("name"))]
        if len(name_matches) == 1:
            return name_matches[0]

    nr_candidates = _seed_by_name_region.get((nn, nr), [])
    blank_vehicle_candidates = [r for r in nr_candidates if not _norm(r.get("vehicle_number"))]
    if len(blank_vehicle_candidates) == 1:
        return blank_vehicle_candidates[0]
    return None


def _seed_for_profile(profile):
    """프로필이 최초 이관된 정확한 Excel 행을 source_row로 다시 찾는다.
    회원명/차량번호가 이후 변경돼도 월별 원장/이월금 계산이 흔들리지 않게 한다.
    """
    _load_seed()
    sr = getattr(profile, "legacy_source_row", None)
    if sr is None:
        return None
    try:
        return (_seed_by_source_row or {}).get(int(sr))
    except Exception:
        return None


def _legacy_reconstructed_balance(profile, month_no: int) -> Optional[int]:
    """2026 기준원장의 특정 월말 잔액을 안전하게 복원한다.

    원본 엑셀에는 일부 월의 ``미수금`` 셀이 비어 있지만 ``청구금`` 또는
    고정 월부과액은 남아 있는 행이 있다. 기존 화면은 이런 달에서 잔액을
    그대로 멈춰 보여 10,000원 부과가 있는데도 잔액이 증가하지 않는 모순이
    생겼다. 명시된 월말 미수금이 있으면 그것을 최우선으로 사용하고, 없으면
    청구금-입금액, 그것도 없으면 전월잔액+고정월부과-입금액 순으로 복원한다.
    실제 DB 잔액이나 입금 저장 로직은 변경하지 않는다.
    """
    try:
        target = max(1, min(12, int(month_no)))
    except Exception:
        return None

    months = list(getattr(profile, "legacy_months", None) or [])
    if not months:
        return int(getattr(profile, "legacy_balance", 0) or 0)

    seed = _seed_for_profile(profile)
    try:
        running = int((seed or {}).get("carryover") or 0)
    except Exception:
        running = 0

    rows = {}
    for row in months:
        try:
            mm = int(row.get("month") or 0)
        except Exception:
            continue
        if 1 <= mm <= 12:
            rows[mm] = row

    touched = bool(running)
    for mm in range(1, target + 1):
        row = rows.get(mm) or {}
        arrears = row.get("arrears")
        billed_total = row.get("billed_total")
        payment = int(row.get("payment") or 0)
        monthly_charge = row.get("monthly_charge")

        if arrears is not None:
            running = int(arrears or 0)
            touched = True
            continue
        if billed_total is not None:
            running = int(billed_total or 0) - payment
            touched = True
            continue
        if monthly_charge is not None and int(monthly_charge or 0) != 0:
            running += int(monthly_charge or 0) - payment
            touched = True
            continue
        if payment:
            running -= payment
            touched = True

    return int(running) if touched else None


def _legacy_balance_as_of(profile, as_of_date: Optional[date]) -> int:
    """폐업자의 기준원장 잔액을 폐업월 시점으로 고정한다."""
    fallback = int(getattr(profile, "legacy_balance", 0) or 0)
    if as_of_date is None or as_of_date.year != 2026:
        return fallback
    reconstructed = _legacy_reconstructed_balance(profile, as_of_date.month)
    return fallback if reconstructed is None else int(reconstructed)


def _infer_account(member) -> str:
    # 가입일자 형식은 26.08.27 같은 2자리 연도도 공식 공통 판정함수에서 가입으로 인정한다.
    return "협회비" if is_association_member(getattr(member, "membership_date", None)) else "관리비"


def _is_bae_vehicle(member) -> bool:
    """2026년 기존 택배 관리비 대상 판정: 차량번호에 한글 '배'가 포함된 차량."""
    return "배" in str(getattr(member, "vehicle_number", "") or "").replace(" ", "")


def _management_first_charge_date(member) -> Optional[date]:
    """관리비 첫 부과일.

    - 2026년: 차량번호에 '배'가 있는 기존 택배 관리비만 기존 기준대로 부과 가능.
    - 비택배 일반 관리비: 2027-01-01 이전에는 절대 자동부과하지 않음.
    - 2027년 이후 신규 비가입자는 기존 자격증명발급일 기준 다음 달 부과 로직을 유지하되
      일반 관리비 시작일(2027-01-01)보다 앞당겨지지 않게 한다.
    """
    first = _business_first_charge_date_raw(member, "관리비")
    if _is_bae_vehicle(member):
        return first
    if first is None:
        return None
    return max(first, GENERAL_MANAGEMENT_START_DATE)


def _billing_basis_date(member, account_type: str) -> Optional[date]:
    """업무 기준 부과기준일.

    - 협회비: 가입일자
    - 관리비: 자격증명발급일자
    - 70세/기타: 별도 자동 신규부과 기준을 만들지 않음

    관련 일자가 없으면 None을 반환한다. 임의로 인가일자/created_at을 대신 쓰지 않는다.
    """
    if account_type == "협회비":
        return _parse_date(getattr(member, "membership_date", None))
    if account_type == "관리비":
        return _parse_date(getattr(member, "certificate_issue_date", None))
    return None


def _business_first_charge_date_raw(member, account_type: str) -> Optional[date]:
    basis = _billing_basis_date(member, account_type)
    return _first_of_next_month(basis) if basis else None


def _business_first_charge_date(member, account_type: str) -> Optional[date]:
    if account_type == "관리비":
        return _management_first_charge_date(member)
    return _business_first_charge_date_raw(member, account_type)


def _make_profile(member, seed=None) -> ReceivableProfile:
    if seed:
        # legacy_balance는 2026-08 원본 장부의 "현재 미수금(8월 미수금)" 그 자체다.
        # 1~8월 부과를 ReceivableCharge로 다시 만들면 이중계산되므로,
        # 프로그램 자동부과는 원장 컷오버 다음 달(2026-09)부터만 시작한다.
        acct = seed.get("account_type") or _infer_account(member)
        return ReceivableProfile(
            member_id=member.id,
            account_type=acct,
            unit_fee=ACCOUNT_FEES.get(acct, 5000),
            vehicle_count=1,
            first_charge_date=LEGACY_NEXT_BILL_DATE.isoformat(),
            legacy_balance=int(seed.get("current_arrears") or 0),
            legacy_months=seed.get("months") or [],
            legacy_source_row=seed.get("source_row"),
            legacy_note=seed.get("legacy_note") or None,
        )

    acct = _infer_account(member)
    first_charge = _business_first_charge_date(member, acct)
    # 신규/비legacy 회원은 협회비=가입일자, 관리비=자격증명발급일자를 기준으로만 부과한다.
    # 기준일이 없으면 first_charge_date=None으로 두어 자동부과하지 않는다.
    # 과거 기존회원이 뒤늦게 profile만 생성된 경우에는 소급부과를 막기 위해 컷오버월까지만 보정한다.
    if first_charge and first_charge < LEGACY_NEXT_BILL_DATE:
        first_charge = LEGACY_NEXT_BILL_DATE
    return ReceivableProfile(
        member_id=member.id,
        account_type=acct,
        unit_fee=ACCOUNT_FEES.get(acct, 5000),
        vehicle_count=1,
        first_charge_date=first_charge.isoformat() if first_charge else None,
        legacy_balance=0,
        legacy_months=[],
    )


def _eligible_missing_profiles_query(db: Session):
    return (
        db.query(models.LicenseHolder)
        .outerjoin(ReceivableProfile, ReceivableProfile.member_id == models.LicenseHolder.id)
        .filter(ReceivableProfile.id.is_(None))
        .filter(or_(models.LicenseHolder.status.is_(None), models.LicenseHolder.status != "pending"))
        .filter(or_(
            models.LicenseHolder.deleted_at.is_(None),
            models.LicenseHolder.closure_id.isnot(None),
            models.LicenseHolder.status != "active",
        ))
        .order_by(models.LicenseHolder.id.desc())
    )


def _sync_missing_profiles_fast(db: Session, limit: int = 100, allow_legacy_seed: bool = False) -> int:
    """신규회원 즉시 노출용. 전체 3천명을 읽지 않고 '프로필 없는 회원'만 인덱스로 조회."""
    members = _eligible_missing_profiles_query(db).limit(limit).all()
    if not members:
        return 0
    created = 0
    for member in members:
        # 관리자 단순삭제(폐업기록 없음)는 미수금 신규대상에서 제외.
        if member.deleted_at is not None and not member.closure_id and (member.status or "active") == "active":
            continue
        db.add(_make_profile(member, _match_seed(member) if allow_legacy_seed else None))
        created += 1
    if created:
        try:
            db.commit()
        except IntegrityError:
            # 여러 직원이 같은 순간 화면을 열어도 unique(member_id)로 중복 방지.
            db.rollback()
    return created


def _sync_profiles_full(db: Session, allow_legacy_seed: bool = False) -> int:
    created = 0
    while True:
        batch = _eligible_missing_profiles_query(db).limit(500).all()
        if not batch:
            break
        batch_created = 0
        for member in batch:
            if member.deleted_at is not None and not member.closure_id and (member.status or "active") == "active":
                # 다시 조회되지 않도록 이 경우는 full sync에서 건너뛰되 loop 탈출 방지
                continue
            db.add(_make_profile(member, _match_seed(member) if allow_legacy_seed else None))
            batch_created += 1
        if not batch_created:
            break
        try:
            db.commit()
            created += batch_created
        except IntegrityError:
            db.rollback()
            # 다른 worker가 동시에 생성한 경우 재조회
    return created


def _baseline_marker(db: Session):
    return db.query(ReceivableSystemState).filter(ReceivableSystemState.key == LEGACY_BASELINE_KEY).first()


def _apply_legacy_baseline_once(db: Session) -> int:
    """최신 미수금 Excel 스냅샷을 DB 공식 원장으로 **최초 1회만** 이관한다.

    이관 완료 여부는 프로세스 메모리가 아니라 receivable_system_state에 영구 기록한다.
    따라서 Railway 재배포/재기동/여러 worker에서도 Excel/JSON이 기존 DB 잔액을 다시
    덮어쓰지 않는다. 이후 수납·선납·부과·폐업·연락은 PostgreSQL 데이터만 누적한다.
    """
    global _legacy_baseline_ready
    if _legacy_baseline_ready:
        return 0

    with _legacy_baseline_lock:
        if _legacy_baseline_ready:
            return 0
        if _baseline_marker(db):
            _legacy_baseline_ready = True
            return 0

        # 최초 이관 때만 기존 회원 profile을 seed와 함께 생성한다.
        _sync_profiles_full(db, allow_legacy_seed=True)

        rows = (
            db.query(ReceivableProfile, models.LicenseHolder)
            .join(models.LicenseHolder, models.LicenseHolder.id == ReceivableProfile.member_id)
            .all()
        )
        imported = 0
        for profile, member in rows:
            seed = _match_seed(member)
            if not seed:
                continue

            profile.legacy_source_row = int(seed.get("source_row")) if seed.get("source_row") is not None else None
            profile.legacy_balance = int(seed.get("current_arrears") or 0)
            profile.legacy_months = seed.get("months") or []
            profile.legacy_note = seed.get("legacy_note") or None
            profile.vehicle_count = 1

            # 기존 원장 계정은 원장 기준. 수동지정 계정만 보존한다.
            if int(getattr(profile, "account_manual_override", 0) or 0) != 1:
                account = seed.get("account_type") or _infer_account(member)
                profile.account_type = account
                profile.unit_fee = ACCOUNT_FEES.get(account, 5000)

            # 2026-08까지의 잔액은 snapshot에 이미 포함되어 있으므로 9월부터 DB 자동부과.
            profile.first_charge_date = LEGACY_NEXT_BILL_DATE.isoformat()
            imported += 1

        payload = json.loads(DATA_FILE.read_text(encoding="utf-8")) if DATA_FILE.exists() else {}
        state_value = json.dumps(
            {
                "source_filename": payload.get("source_filename", ""),
                "source_sha256": payload.get("source_sha256", ""),
                "source_rows": len(payload.get("rows", []) or []),
                "matched_profiles": imported,
                "cutover_through": LEGACY_DATA_THROUGH_KEY,
                "next_bill_date": LEGACY_NEXT_BILL_DATE.isoformat(),
                "ledger_mode": LEDGER_MODE,
                "excel_reupload_required": False,
                "applied_at": datetime.now(KST).isoformat(),
            },
            ensure_ascii=False,
        )
        db.add(ReceivableSystemState(key=LEGACY_BASELINE_KEY, value=state_value))
        try:
            db.commit()
        except IntegrityError:
            # 다중 worker가 동시에 최초 이관을 시도한 경우 PK marker가 중복을 차단한다.
            db.rollback()
            if not _baseline_marker(db):
                raise
        _legacy_baseline_ready = True
        return imported


def _refresh_legacy_baseline_if_seed_changed(db: Session, marker) -> int:
    """번들된 최신 기준원장이 기존 DB baseline과 다를 때 **한 번만** 갱신한다.

    수납/선납/연락은 별도 테이블이므로 절대 삭제하지 않는다. 이 함수가 바꾸는 것은
    2026-08 기준 snapshot(legacy_balance/legacy_months/계정/대수)뿐이다.
    marker에 최신 source_sha256을 기록하므로 같은 파일로 재배포해도 다시 덮어쓰지 않는다.
    """
    if marker is None or not DATA_FILE.exists():
        return 0
    try:
        payload = json.loads(DATA_FILE.read_text(encoding="utf-8"))
    except Exception:
        return 0
    current_sha = str(payload.get("source_sha256") or "").strip()
    if not current_sha:
        return 0
    try:
        info = json.loads(marker.value or "{}") if marker.value else {}
    except Exception:
        info = {}
    applied_sha = str(info.get("source_sha256") or "").strip()
    if applied_sha == current_sha:
        return 0

    _load_seed()
    rows = (
        db.query(ReceivableProfile, models.LicenseHolder)
        .join(models.LicenseHolder, models.LicenseHolder.id == ReceivableProfile.member_id)
        .all()
    )
    refreshed = 0
    for profile, member in rows:
        seed = _seed_for_profile(profile) or _match_seed(member)
        if not seed:
            continue
        profile.legacy_source_row = int(seed.get("source_row")) if seed.get("source_row") is not None else None
        profile.legacy_balance = int(seed.get("current_arrears") or 0)
        profile.legacy_months = seed.get("months") or []
        profile.legacy_note = seed.get("legacy_note") or None
        profile.vehicle_count = 1
        if int(getattr(profile, "account_manual_override", 0) or 0) != 1:
            account = seed.get("account_type") or _infer_account(member)
            profile.account_type = account
            profile.unit_fee = ACCOUNT_FEES.get(account, 5000)
        profile.first_charge_date = LEGACY_NEXT_BILL_DATE.isoformat()
        refreshed += 1

    info.update({
        "source_filename": payload.get("source_filename", ""),
        "source_sha256": current_sha,
        "source_rows": len(payload.get("rows", []) or []),
        "matched_profiles": refreshed,
        "cutover_through": LEGACY_DATA_THROUGH_KEY,
        "next_bill_date": LEGACY_NEXT_BILL_DATE.isoformat(),
        "ledger_mode": LEDGER_MODE,
        "excel_reupload_required": False,
        "baseline_refreshed_at": datetime.now(KST).isoformat(),
        "baseline_refresh_reason": "bundled_final_snapshot_changed",
    })
    marker.value = json.dumps(info, ensure_ascii=False)
    db.commit()
    return refreshed


def _repair_non_receivable_import_payments_once(db: Session) -> int:
    """과거 버그로 일괄수납된 자격증명/발급비 등 비미수금 거래를 1회 안전 취소한다.

    ReceivableImportRow 원본에 강한 키워드 증거가 있고 payment_id로 정확히 연결된 건만 취소한다.
    수동입금이나 근거가 불명확한 거래는 건드리지 않는다.
    """
    marker = db.query(ReceivableSystemState).filter(ReceivableSystemState.key == NON_RECEIVABLE_REPAIR_KEY).first()
    if marker is not None:
        return 0

    rows = (
        db.query(ReceivableImportRow)
        .filter(ReceivableImportRow.status == "posted", ReceivableImportRow.payment_id.isnot(None))
        .all()
    )
    repaired = 0
    batch_ids = set()
    now = datetime.now(KST)
    for row in rows:
        reason = _non_receivable_import_reason({
            "payer_name": row.payer_name or "",
            "memo": row.memo or "",
            "external_id": row.external_id or "",
            "raw_data": row.raw_data or {},
        })
        if not reason:
            continue
        payment = (
            db.query(ReceivablePayment)
            .filter(ReceivablePayment.id == row.payment_id, ReceivablePayment.cancelled_at.is_(None))
            .first()
        )
        if payment is None:
            continue
        payment.cancelled_at = now
        payment.cancelled_by = "system:nonreceivable-20260831"
        row.status = "review"
        row.match_reason = f"{reason} · 과거 자동반영 취소 · 재확인 필요"
        batch_ids.add(row.batch_id)
        repaired += 1

    for batch_id in batch_ids:
        batch = db.query(ReceivableImportBatch).filter(ReceivableImportBatch.id == batch_id).first()
        if batch is not None:
            _refresh_import_batch(db, batch)
            batch.status = "partial" if batch.review_rows else batch.status

    db.add(ReceivableSystemState(
        key=NON_RECEIVABLE_REPAIR_KEY,
        value=json.dumps({
            "repaired": repaired,
            "rule": "certificate/issuance/vehicle-replacement/join-fee imports are review-only",
            "applied_at": now.isoformat(),
        }, ensure_ascii=False),
    ))
    db.commit()
    return repaired


def _repair_fixed_monthly_fees_once(db: Session) -> dict:
    """기존 DB에 남아 있는 대수배수/오금액 자동부과를 배포 후 1회 정리한다."""
    marker = db.query(ReceivableSystemState).filter(ReceivableSystemState.key == FEE_REPAIR_KEY).first()
    if marker is not None:
        return {"profiles_fixed": 0, "charges_removed": 0, "charges_created": 0}
    profiles_fixed = _repair_profile_fee_rules(db)
    charges_removed = _repair_invalid_auto_charges(db)
    charges_created = _sync_charges(db) if charges_removed else 0
    db.add(ReceivableSystemState(
        key=FEE_REPAIR_KEY,
        value=json.dumps({
            "rules": ACCOUNT_FEES,
            "vehicle_multiplier": False,
            "profiles_fixed": profiles_fixed,
            "charges_removed": charges_removed,
            "charges_created": charges_created,
            "applied_at": datetime.now(KST).isoformat(),
        }, ensure_ascii=False),
    ))
    db.commit()
    return {"profiles_fixed": profiles_fixed, "charges_removed": charges_removed, "charges_created": charges_created}


def _repair_general_management_2027_rule_once(db: Session) -> dict:
    """2026년에 잘못 생성된 비택배 일반 관리비 자동부과를 1회 제거한다.

    원본 2026 legacy 잔액/월별장부, 실제 입금, 수동조정은 건드리지 않는다.
    source='auto'인 잘못된 부과만 삭제하고 비택배 관리비 첫 부과일을 2027-01 이후로 보정한다.
    """
    marker = db.query(ReceivableSystemState).filter(
        ReceivableSystemState.key == MANAGEMENT_2027_REPAIR_KEY
    ).first()
    if marker is not None:
        return {"profiles_fixed": 0, "charges_removed": 0}

    rows = (
        db.query(ReceivableProfile, models.LicenseHolder)
        .join(models.LicenseHolder, models.LicenseHolder.id == ReceivableProfile.member_id)
        .filter(ReceivableProfile.account_type == "관리비")
        .all()
    )
    profiles_fixed = 0
    for profile, member in rows:
        if _is_bae_vehicle(member):
            continue
        first = _business_first_charge_date(member, "관리비")
        target = first.isoformat() if first else None
        if (profile.first_charge_date or None) != target:
            profile.first_charge_date = target
            profiles_fixed += 1
    if profiles_fixed:
        db.commit()

    charges_removed = _repair_invalid_auto_charges(db)
    db.add(ReceivableSystemState(
        key=MANAGEMENT_2027_REPAIR_KEY,
        value=json.dumps({
            "rule": "2026 non-bae management fee disabled; general management starts 2027-01",
            "profiles_fixed": profiles_fixed,
            "charges_removed": charges_removed,
            "applied_at": datetime.now(KST).isoformat(),
        }, ensure_ascii=False),
    ))
    db.commit()
    return {"profiles_fixed": profiles_fixed, "charges_removed": charges_removed}


def _ensure_db_ledger_ready(db: Session) -> int:
    """DB 공식원장 준비.

    1) receivables 전용 테이블을 lazy-create하여 Railway background migration보다
       사용자가 먼저 화면을 열어도 500/0명 화면이 나오지 않게 한다.
    2) marker만 남고 실제 profile/legacy snapshot이 비어 있는 불완전 이관 상태는
       자동 복구한다. 기존 payments/contact_logs는 삭제하거나 덮어쓰지 않는다.
    """
    _ensure_receivables_schema_ready()
    # 기존 DB에 잘못 반영된 자격증명/발급비 일괄수납은 증거가 확실한 건만 1회 취소한다.
    _repair_non_receivable_import_payments_once(db)
    # 협회비 10,000 / 관리비·70세 5,000 외의 잘못된 자동부과도 1회 정리한다.
    _repair_fixed_monthly_fees_once(db)
    # 2026 비택배 일반 관리비 오부과를 제거하고 2027-01 시작 규칙을 고정한다.
    _repair_general_management_2027_rule_once(db)

    marker = _baseline_marker(db)
    if marker is not None:
        # 최종 기준원장 파일이 바뀐 경우에만 baseline snapshot을 1회 갱신한다.
        # 프로그램에서 입력한 payments/contact_logs는 그대로 보존된다.
        _refresh_legacy_baseline_if_seed_changed(db, marker)
        profile_count = db.query(func.count(ReceivableProfile.id)).scalar() or 0
        legacy_count = (
            db.query(func.count(ReceivableProfile.id))
            .filter(ReceivableProfile.legacy_source_row.isnot(None))
            .scalar()
            or 0
        )

        # 과거 배포 중 marker만 생성됐거나 profile이 비어버린 경우 self-heal.
        # 최신 Excel snapshot은 이 시점에만 baseline 필드로 복구되고,
        # 이후 프로그램에서 쌓인 입금/선납/연락 행은 별도 테이블이므로 그대로 보존된다.
        if profile_count == 0 or legacy_count == 0:
            _sync_profiles_full(db, allow_legacy_seed=True)
            rows = (
                db.query(ReceivableProfile, models.LicenseHolder)
                .join(models.LicenseHolder, models.LicenseHolder.id == ReceivableProfile.member_id)
                .all()
            )
            repaired = 0
            for profile, member in rows:
                seed = _match_seed(member)
                if not seed:
                    continue
                profile.legacy_source_row = int(seed.get('source_row')) if seed.get('source_row') is not None else None
                profile.legacy_balance = int(seed.get('current_arrears') or 0)
                profile.legacy_months = seed.get('months') or []
                profile.legacy_note = seed.get('legacy_note') or None
                profile.vehicle_count = 1
                if int(getattr(profile, 'account_manual_override', 0) or 0) != 1:
                    account = seed.get('account_type') or _infer_account(member)
                    profile.account_type = account
                    profile.unit_fee = ACCOUNT_FEES.get(account, 5000)
                profile.first_charge_date = LEGACY_NEXT_BILL_DATE.isoformat()
                repaired += 1
            db.commit()

            # marker value도 실제 복구 결과로 갱신하여 다음 재기동에서 재작업하지 않는다.
            try:
                info = json.loads(marker.value or '{}') if marker.value else {}
            except Exception:
                info = {}
            info.update({
                'matched_profiles': repaired,
                'recovered_at': datetime.now(KST).isoformat(),
                'recovery_reason': 'marker_without_legacy_profiles',
                'ledger_mode': LEDGER_MODE,
                'excel_reupload_required': False,
            })
            marker.value = json.dumps(info, ensure_ascii=False)
            db.commit()

            global _legacy_baseline_ready
            _legacy_baseline_ready = True
            return repaired

    return _apply_legacy_baseline_once(db)


def _repair_account_types(db: Session) -> int:
    """legacy/수동계정은 보존하고, 나머지만 가입일자 기준으로 정합성 보정."""
    rows = (
        db.query(ReceivableProfile, models.LicenseHolder)
        .join(models.LicenseHolder, models.LicenseHolder.id == ReceivableProfile.member_id)
        .filter(ReceivableProfile.legacy_source_row.is_(None))
        .filter(func.coalesce(ReceivableProfile.account_manual_override, 0) != 1)
        .all()
    )
    fixed = 0
    for p, member in rows:
        correct = _infer_account(member)
        changed = False
        if p.account_type != correct or int(p.unit_fee or 0) != ACCOUNT_FEES.get(correct, 5000):
            p.account_type = correct
            p.unit_fee = ACCOUNT_FEES.get(correct, 5000)
            changed = True
        if int(p.vehicle_count or 1) != 1:
            p.vehicle_count = 1
            changed = True

        # 비legacy 자동계정은 부과기준일도 현재 회원정보와 항상 맞춘다.
        # 협회비=가입일자 다음달 1일, 관리비=자격증명발급일자 다음달 1일, 없으면 미부과(None).
        first = _business_first_charge_date(member, correct)
        if first and first < LEGACY_NEXT_BILL_DATE:
            first = LEGACY_NEXT_BILL_DATE
        target_first = first.isoformat() if first else None
        if (p.first_charge_date or None) != target_first:
            p.first_charge_date = target_first
            changed = True

        if changed:
            fixed += 1
    if fixed:
        db.commit()
    return fixed


def _closure_map(db: Session):
    """현재 폐업 상태인 회원의 *현재 폐업건*만 member_id -> Closure로 반환한다.

    과거 폐업/양도/이관 이력이 같은 member_id에 남아 있어도 현재 회원이 active이면
    자동부과 중단/폐업미수 판정에 사용하면 안 된다. 현재 상태의 기준은
    license_holders.status + license_holders.closure_id 이다.

    구자료 중 status='closed'인데 closure_id가 비어 있는 경우에만 최신 member_id 폐업건을
    보조적으로 사용한다. active 회원에는 절대 과거 폐업건을 붙이지 않는다.
    """
    members = (
        db.query(models.LicenseHolder)
        .filter(
            models.LicenseHolder.deleted_at.is_(None),
            models.LicenseHolder.status == "closed",
        )
        .all()
    )
    if not members:
        return {}

    by_closure_id = {}
    explicit_ids = {m.closure_id for m in members if getattr(m, "closure_id", None)}
    if explicit_ids:
        for c in (
            db.query(models.Closure)
            .filter(
                models.Closure.id.in_(explicit_ids),
                models.Closure.deleted_at.is_(None),
            )
            .all()
        ):
            by_closure_id[c.id] = c

    out = {}
    missing = []
    for m in members:
        cid = getattr(m, "closure_id", None)
        c = by_closure_id.get(cid) if cid else None
        if c is not None:
            out[m.id] = c
        else:
            missing.append(m.id)

    # 하위호환: status=closed이지만 closure_id가 없는 구자료만 최신 연결 폐업건으로 보강.
    if missing:
        rows = (
            db.query(models.Closure)
            .filter(
                models.Closure.member_id.in_(missing),
                models.Closure.deleted_at.is_(None),
            )
            .order_by(models.Closure.id.desc())
            .all()
        )
        for c in rows:
            if c.member_id and c.member_id not in out:
                out[c.member_id] = c
    return out


def _month_key(d: date) -> str:
    return f"{d.year:04d}-{d.month:02d}"


def _has_legacy_evidence(member, profile) -> bool:
    """기존 MUSTARD/2026 원장 회원인지 DB 상태와 원본 seed를 함께 보고 판정한다.

    과거 버전에서 legacy_source_row가 비어 저장된 profile이 존재할 수 있으므로
    그 컬럼 하나만 믿으면 기존회원 수천 명이 신규 부과대기로 오판된다.
    """
    if getattr(profile, "legacy_source_row", None) is not None:
        return True

    months = getattr(profile, "legacy_months", None) or []
    if months:
        for row in months:
            if not isinstance(row, dict):
                continue
            if (
                row.get("billed_total") is not None
                or row.get("payment") is not None
                or row.get("arrears") is not None
                or bool(row.get("payment_date"))
            ):
                return True

    # legacy_balance가 0이 아닌 기존회원도 명백한 원장 이관회원이다.
    if int(getattr(profile, "legacy_balance", 0) or 0) != 0:
        return True

    # 컷오버 이후 신규 판정에서 bundled Excel/JSON을 다시 조회하지 않는다.
    # 최초 이관 때 매칭된 회원은 legacy_source_row/legacy_months/legacy_balance에 영구 표식이 남는다.
    return False


def _true_registration_date(member, profile=None) -> Optional[date]:
    """신규 부과 판정용 실제 업무 기준일.

    협회비는 가입일자, 관리비는 자격증명발급일자만 사용한다.
    created_at/인가일자는 부과 기준으로 사용하지 않는다.
    """
    account_type = getattr(profile, "account_type", None) or _infer_account(member)
    return _billing_basis_date(member, account_type)


def _is_true_new_member(member, profile) -> bool:
    """'부과대기'는 원장에 없고 실제 업무 기준일이 컷오버 이후인 회원에게만 붙인다."""
    if _has_legacy_evidence(member, profile):
        return False
    reg = _true_registration_date(member, profile)
    return bool(reg and reg >= LEGACY_NEW_MEMBER_CUTOFF)


def _repair_profile_fee_rules(db: Session) -> int:
    """모든 미수금 프로필의 월회비를 계정별 고정액으로 강제한다.

    협회비 10,000 / 관리비 5,000 / 70세 5,000. 차량 대수로 곱하지 않는다.
    과거 UI에서 20,000/40,000원처럼 보이거나 잘못 생성된 자동부과의 재발을 막는다.
    """
    fixed = 0
    for p in db.query(ReceivableProfile).all():
        expected = ACCOUNT_FEES.get(p.account_type)
        if expected is None:
            continue
        changed = False
        if int(p.unit_fee or 0) != expected:
            p.unit_fee = expected
            changed = True
        if int(p.vehicle_count or 1) != 1:
            p.vehicle_count = 1
            changed = True
        if changed:
            fixed += 1
    if fixed:
        db.commit()
    return fixed


def _auto_charge_allowed_for_month(profile, member, billing_month: str) -> bool:
    """계정/차량/시행연도 기준 자동부과 허용 여부.

    핵심 업무규칙:
    - 2026-12까지 관리비는 차량번호에 '배'가 포함된 기존 택배 관리비만 허용.
    - 비택배 일반 관리비는 2027-01부터 허용.
    - 협회비/70세는 이 함수에서 추가 제한하지 않음.
    """
    if str(getattr(profile, "account_type", "") or "") != "관리비":
        return True
    if str(billing_month or "") < GENERAL_MANAGEMENT_START_KEY:
        return _is_bae_vehicle(member)
    return True


def _valid_auto_charge(profile, member, closure, charge, today: Optional[date] = None) -> bool:
    """현재 미수금에 포함해도 되는 자동부과만 판정한다.

    - 미래월 자동부과: 무효
    - legacy 회원 2026-08 이전/당월 자동부과: 원본 원장과 중복이므로 무효
    - 폐업/양도/이관 월 이후 자동부과: 무효
    """
    today = today or datetime.now(KST).date()
    month = str(charge.billing_month or "")
    if not month or month > _month_key(today):
        return False
    if _has_legacy_evidence(member, profile) and month <= LEGACY_DATA_THROUGH_KEY:
        return False
    if not _auto_charge_allowed_for_month(profile, member, month):
        return False
    if closure is not None:
        close_d = _parse_date(getattr(closure, "closure_date", None))
        if close_d and month > _month_key(close_d):
            return False
    expected = ACCOUNT_FEES.get(profile.account_type)
    if expected is None:
        return False
    if str(charge.account_type or "") != str(profile.account_type or ""):
        return False
    if int(charge.amount or 0) != int(expected):
        return False
    return True


def _repair_invalid_auto_charges(db: Session) -> int:
    """이전 버그로 생성된 잘못된 auto charge만 삭제한다.
    원본 legacy 장부/입금/수동 데이터는 절대 건드리지 않는다.
    """
    rows = db.query(ReceivableCharge).filter(ReceivableCharge.source == "auto").all()
    if not rows:
        return 0
    profiles = {p.member_id: p for p in db.query(ReceivableProfile).all()}
    members = {m.id: m for m in db.query(models.LicenseHolder).all()}
    closures = _closure_map(db)
    today = datetime.now(KST).date()
    removed = 0
    for ch in rows:
        p = profiles.get(ch.member_id)
        m = members.get(ch.member_id)
        if not p or not m or not _valid_auto_charge(p, m, closures.get(ch.member_id), ch, today):
            db.delete(ch)
            removed += 1
    if removed:
        db.commit()
    return removed


def _sync_charges(db: Session) -> int:
    """월 자동부과 생성. 수동 sync/백그라운드에서만 실행되어 화면 조회를 막지 않는다."""
    today = datetime.now(KST).date()
    members = {m.id: m for m in db.query(models.LicenseHolder).all()}
    closures = _closure_map(db)
    existing = {
        (mid, month)
        for mid, month in db.query(ReceivableCharge.member_id, ReceivableCharge.billing_month).all()
    }
    added = 0
    for p in db.query(ReceivableProfile).all():
        member = members.get(p.member_id)
        if not member or not p.first_charge_date:
            continue
        start = _parse_date(p.first_charge_date)
        if not start:
            continue
        end = today
        cl = closures.get(member.id)
        if cl is not None:
            close_d = _parse_date(getattr(cl, "closure_date", None)) or today
            end = close_d
        elif getattr(member, "deleted_at", None) is not None or (getattr(member, "status", None) or "") == "pending":
            # 실제 Closure 없이 관리자 삭제/예정 상태인 행에는 자동부과하지 않는다.
            continue
        if date(start.year, start.month, 1) > date(end.year, end.month, 1):
            continue
        for month_start in _month_iter(start, end):
            month_key = month_start.strftime("%Y-%m")
            if not _auto_charge_allowed_for_month(p, member, month_key):
                continue
            key = (p.member_id, month_key)
            if key in existing:
                continue
            db.add(
                ReceivableCharge(
                    member_id=p.member_id,
                    billing_month=key[1],
                    amount=int(ACCOUNT_FEES.get(p.account_type, 0)),
                    account_type=p.account_type,
                    source="auto",
                )
            )
            existing.add(key)
            added += 1
    if added:
        try:
            db.commit()
        except IntegrityError:
            db.rollback()
            # UNIQUE(member_id,billing_month)이 최종적으로 중복부과를 차단한다.
    return added


def _ensure_current_month_billing(db: Session) -> int:
    """현재 월 자동부과를 월 1회 보장한다.

    2026-08까지는 최신 원장 snapshot에 포함되어 있어 자동부과하지 않는다.
    2026-09부터는 Railway 백그라운드 스케줄러가 월 변경을 감지해 자동 생성하며,
    화면 최초 접근/수동 동기화도 안전망으로 동일 함수를 호출한다. 완료 marker를 영구 저장해
    재기동이나 중복 호출에도 같은 월을 두 번 부과하지 않는다.
    """
    global _billing_ready_month
    today = datetime.now(KST).date()
    month_key = _month_key(today)
    if month_key <= LEGACY_DATA_THROUGH_KEY:
        _billing_ready_month = month_key
        return 0
    if _billing_ready_month == month_key:
        return 0

    state_key = f"receivable_billing_generated:{month_key}"
    with _billing_month_lock:
        if _billing_ready_month == month_key:
            return 0
        if db.query(ReceivableSystemState).filter(ReceivableSystemState.key == state_key).first():
            _billing_ready_month = month_key
            return 0

        # 부과 직전 계정/폐업 상태와 고정월회비를 다시 확인한다.
        _repair_account_types(db)
        _repair_profile_fee_rules(db)
        _repair_invalid_auto_charges(db)
        added = _sync_charges(db)
        db.add(
            ReceivableSystemState(
                key=state_key,
                value=json.dumps(
                    {"billing_month": month_key, "charges_created": added, "applied_at": datetime.now(KST).isoformat()},
                    ensure_ascii=False,
                ),
            )
        )
        try:
            db.commit()
        except IntegrityError:
            db.rollback()
            # 다른 worker가 먼저 marker를 만들었을 수 있다.
            if not db.query(ReceivableSystemState).filter(ReceivableSystemState.key == state_key).first():
                raise
        _billing_ready_month = month_key
        return added


def _sync_all(db: Session):
    """운영 DB 동기화. Excel/JSON 재반영은 절대 하지 않는다."""
    baseline_imported = _ensure_db_ledger_ready(db)
    p = _sync_profiles_full(db, allow_legacy_seed=False)
    r = _repair_account_types(db)
    fee_fixed = _repair_profile_fee_rules(db)
    removed = _repair_invalid_auto_charges(db)
    c = _ensure_current_month_billing(db)
    # 신규 profile 또는 가입일자/자격증명발급일자 변경으로 이번 달부터 부과대상이 된 경우 즉시 보정.
    if p or r or fee_fixed or removed:
        c += _sync_charges(db)
    return {
        "profiles_created": p,
        "charges_created": c,
        "charges_removed": removed,
        "accounts_repaired": r,
        "fee_profiles_repaired": fee_fixed,
        "legacy_baseline_imported_once": baseline_imported,
        "ledger_mode": LEDGER_MODE,
        "excel_reupload_required": False,
    }


def _schedule_background_sync(background_tasks: BackgroundTasks):
    """호환용 no-op.

    과거에는 목록/요약 요청마다 10분 간격으로 3천명 전체 sync를 백그라운드에서 돌렸다.
    운영 DB가 공식원장이 된 뒤에는 그 전수작업 자체를 없애 DB 경합과 지연을 줄인다.
    신규회원은 목록 요청의 missing-profile 쿼리로 즉시 연결되고, 월 부과는 /sync 및
    수납화면 최초 접근 시 필요한 범위에서 보장한다.
    """
    return None

def _monthly_billing_worker():
    """Railway 서비스가 살아 있는 동안 월 변경을 자동 감지해 월 부과를 생성한다.

    직원이 화면을 열지 않아도 동작한다. 동일 회원/동일 월은 DB UNIQUE 제약과
    ReceivableSystemState marker가 이중부과를 차단한다.
    """
    while True:
        db = SessionLocal()
        try:
            _ensure_receivables_schema_ready()
            _ensure_db_ledger_ready(db)
            # 프로필이 없는 신규회원만 빠르게 연결하고, 비legacy 회원의 계정/첫부과 기준일만 정합성 보정한다.
            created = _sync_missing_profiles_fast(db, limit=500, allow_legacy_seed=False)
            repaired = _repair_account_types(db)
            _ensure_current_month_billing(db)
            if created or repaired:
                _sync_charges(db)
        except Exception as exc:
            print(f"[receivables monthly scheduler] {type(exc).__name__}: {exc}")
            try:
                db.rollback()
            except Exception:
                pass
        finally:
            db.close()
        # 월 경계 누락 방지를 위해 15분 간격 확인. 실제 전수 부과는 월 1회만 실행된다.
        time.sleep(900)


@router.on_event("startup")
def _start_monthly_billing_scheduler():
    global _monthly_scheduler_started
    with _monthly_scheduler_lock:
        if _monthly_scheduler_started:
            return
        _monthly_scheduler_started = True
        threading.Thread(
            target=_monthly_billing_worker,
            name="receivables-monthly-billing",
            daemon=True,
        ).start()


def _legacy_snapshot_cutoff_at() -> datetime:
    """최신 통장기준 원장이 확정된 시각.

    legacy 회원은 이 시각 이전에 DB에 이미 존재하던 2026-08-31 이하 수납/잔액수정을
    최신 원장에 포함된 과거 처리로 본다. 기록 자체는 삭제하지 않고 잔액 계산에서만
    제외한다. 이 시각 이후 새로 입력한 수납은 같은 8/31 날짜라도 정상 반영한다.
    """
    fallback = datetime(2026, 8, 31, 20, 13, 23, tzinfo=KST)
    try:
        if DATA_FILE.exists():
            payload = json.loads(DATA_FILE.read_text(encoding="utf-8"))
            raw = str(payload.get("source_cutoff_at_kst") or "").strip()
            if raw:
                dt = datetime.fromisoformat(raw)
                if dt.tzinfo is None:
                    dt = dt.replace(tzinfo=KST)
                return dt
    except Exception:
        pass
    return fallback


def _legacy_payment_is_snapshot_covered(profile: ReceivableProfile, payment: ReceivablePayment) -> bool:
    """최신 8/31 원장에 이미 흡수된 과거 DB 수납인지 판정한다.

    - legacy 회원만 대상
    - 지급/수정 기준일이 2026-08-31 이하
    - DB 생성시각이 최신 기준원장 확정시각 이하
    위 세 조건을 모두 만족할 때만 잔액에서 중복 차감하지 않는다.
    """
    if getattr(profile, "legacy_source_row", None) is None:
        return False
    pd = str(getattr(payment, "payment_date", None) or "")[:10]
    if not pd or pd > LEGACY_DATA_THROUGH_DATE_ISO:
        return False
    created = getattr(payment, "created_at", None)
    if created is None:
        # 과거 데이터 중 created_at이 없는 건은 최신 snapshot 이전 기록으로 취급한다.
        return True
    try:
        cutoff = _legacy_snapshot_cutoff_at()
        if created.tzinfo is None:
            created = created.replace(tzinfo=KST)
        return created <= cutoff
    except Exception:
        return True


def _legacy_effective_payment_sql_condition():
    """ReceivableProfile을 join한 쿼리에서 사용하는 중복차감 방지 조건."""
    cutoff = _legacy_snapshot_cutoff_at()
    return or_(
        ReceivableProfile.legacy_source_row.is_(None),
        ReceivablePayment.payment_date > LEGACY_DATA_THROUGH_DATE_ISO,
        ReceivablePayment.created_at > cutoff,
    )


def _charge_payment_subqueries(db: Session):
    """잔액 집계용 서브쿼리.

    legacy_balance가 최신 2026-08 통장기준 원장의 누적 미수이므로, legacy 회원의
    1~8월 auto charge를 제외한다. 또한 최신 snapshot 확정 이전에 DB에 이미 존재하던
    8/31 이하 수납/잔액수정은 snapshot에 흡수된 과거 처리이므로 다시 차감하지 않는다.
    기존 DB 기록은 삭제하지 않으며 snapshot 이후 새 입력은 정상 반영한다.
    """
    today_month = _month_key(datetime.now(KST).date())
    charges_sq = (
        db.query(
            ReceivableCharge.member_id.label("member_id"),
            func.coalesce(func.sum(ReceivableCharge.amount), 0).label("charge_total"),
        )
        .join(ReceivableProfile, ReceivableProfile.member_id == ReceivableCharge.member_id)
        .filter(ReceivableCharge.billing_month <= today_month)
        .filter(or_(
            ReceivableProfile.legacy_source_row.is_(None),
            ReceivableCharge.billing_month > LEGACY_DATA_THROUGH_KEY,
        ))
        .group_by(ReceivableCharge.member_id)
        .subquery()
    )
    payments_sq = (
        db.query(
            ReceivablePayment.member_id.label("member_id"),
            func.coalesce(func.sum(ReceivablePayment.amount), 0).label("payment_total"),
        )
        .join(ReceivableProfile, ReceivableProfile.member_id == ReceivablePayment.member_id)
        .filter(ReceivablePayment.cancelled_at.is_(None))
        .filter(_legacy_effective_payment_sql_condition())
        .group_by(ReceivablePayment.member_id)
        .subquery()
    )
    return charges_sq, payments_sq


def _latest_contact_subquery(db: Session):
    ranked = (
        db.query(
            ReceivableContactLog.member_id.label("member_id"),
            ReceivableContactLog.status.label("status"),
            ReceivableContactLog.contact_date.label("contact_date"),
            func.row_number().over(
                partition_by=ReceivableContactLog.member_id,
                order_by=(ReceivableContactLog.contact_date.desc(), ReceivableContactLog.id.desc()),
            ).label("rn"),
        )
        .subquery()
    )
    return (
        db.query(
            ranked.c.member_id.label("member_id"),
            ranked.c.status.label("status"),
            ranked.c.contact_date.label("contact_date"),
        )
        .filter(ranked.c.rn == 1)
        .subquery()
    )


def _current_closure_subquery(db: Session):
    """LicenseHolder.closure_id가 가리키는 현재 폐업건의 메타데이터.

    과거 Closure.member_id 이력은 현재 활성/폐업 판정에 사용하지 않는다.
    """
    return (
        db.query(
            models.Closure.id.label("closure_id"),
            models.Closure.member_id.label("member_id"),
            models.Closure.management_number.label("management_number"),
            models.Closure.closure_date.label("closure_date"),
            models.Closure.closure_type.label("closure_type"),
            models.Closure.reason.label("reason"),
            models.Closure.transferee.label("transferee"),
            models.Closure.transfer_region.label("transfer_region"),
            models.Closure.receipt_date.label("receipt_date"),
            models.Closure.data_type.label("data_type"),
        )
        .filter(models.Closure.deleted_at.is_(None))
        .subquery()
    )


def _receivable_active_sql(_current_closure_sq=None):
    """현재 회원관리의 status를 그대로 따른 활성회원 판정."""
    return and_(
        models.LicenseHolder.deleted_at.is_(None),
        or_(models.LicenseHolder.status.is_(None), models.LicenseHolder.status == "active"),
    )


def _receivable_closed_sql(_current_closure_sq=None):
    """현재 회원관리에서 실제 closed 상태인 회원만 폐업 현재회원으로 판정."""
    return and_(
        models.LicenseHolder.deleted_at.is_(None),
        models.LicenseHolder.status == "closed",
    )


def _billing_state(member, profile, balance: int, is_closed: bool = False) -> str:
    today = datetime.now(KST).date()
    first = _parse_date(profile.first_charge_date)
    if is_closed:
        if balance > 0:
            return "폐업 미수"
        if balance < 0:
            return "폐업 선납"
        return "폐업 완납"
    if not _has_legacy_evidence(member, profile) and not first:
        return "부과기준일 없음"
    # 핵심: 기존 MUSTARD/엑셀 원장 회원은 first_charge_date가 9/1이어도 신규가 아니다.
    if _is_true_new_member(member, profile) and first and first > today:
        return "부과대기"
    if balance > 0:
        return "미수"
    if balance < 0:
        return "선납"
    return "완납"


def _serialize_member(
    member,
    profile,
    balance,
    contact_status="미연락",
    last_contact_date="",
    closure_id=None,
    closure_management_number="",
    closure_date="",
    closure_type="",
    closure_reason="",
    transferee="",
    transfer_region="",
    closure_receipt_date="",
):
    is_closed = (getattr(member, "status", None) or "active") == "closed"
    canonical_membership = "가입" if is_association_member(getattr(member, "membership_date", None)) else "미가입"
    billing_state = _billing_state(member, profile, int(balance), is_closed=is_closed)
    # 화면의 첫 부과일은 내부 DB 컷오버일이 아니라 실제 업무 기준일로 표시한다.
    # 협회비=가입일자 다음달 1일, 관리비=자격증명발급일자 다음달 1일, 기준일 없으면 0.
    business_first = _business_first_charge_date(member, profile.account_type)
    display_first_charge = business_first.isoformat() if business_first else "0"
    return {
        "member_id": member.id,
        "name": member.name or "",
        "management_number": getattr(member, "management_number", None) or "",
        "vehicle_number": member.vehicle_number or "",
        "region": member.region or "",
        "address": getattr(member, "address", None) or "",
        "mobile": getattr(member, "mobile", None) or "",
        "phone": getattr(member, "phone", None) or "",
        "category": member.category or "",
        "account_type": profile.account_type,
        "unit_fee": int(profile.unit_fee or 0),
        "vehicle_count": int(profile.vehicle_count or 1),
        "membership_status": canonical_membership,
        "membership_date": member.membership_date or "",
        "certificate_issue_date": getattr(member, "certificate_issue_date", None) or "",
        "member_status": (closure_type or "폐업") if is_closed else "활성",
        "active": not is_closed,
        "balance": int(balance),
        "arrears_amount": max(int(balance), 0),
        "prepaid_amount": max(-int(balance), 0),
        "first_charge_date": display_first_charge or "",
        "billing_state": billing_state,
        "legacy_member": profile.legacy_source_row is not None,
        "closure_id": closure_id,
        "closure_management_number": closure_management_number or "",
        "closure_date": closure_date or "",
        "closure_type": closure_type or "",
        "closure_reason": closure_reason or "",
        "transferee": transferee or "",
        "transfer_region": transfer_region or "",
        "closure_receipt_date": closure_receipt_date or "",
        "contact_status": contact_status or "미연락",
        "last_contact_date": last_contact_date or "",
    }


def _ensure_profile_for_member(db: Session, member_id: int):
    p = db.query(ReceivableProfile).filter(ReceivableProfile.member_id == member_id).first()
    if p:
        return p
    member = db.query(models.LicenseHolder).filter(models.LicenseHolder.id == member_id).first()
    if not member:
        return None
    if (member.status or "active") == "pending":
        return None
    _ensure_db_ledger_ready(db)
    p = _make_profile(member, None)
    db.add(p)
    try:
        db.commit()
        db.refresh(p)
    except IntegrityError:
        db.rollback()
        p = db.query(ReceivableProfile).filter(ReceivableProfile.member_id == member_id).first()
    return p


@router.get("/receivables")
def receivables_page():
    return FileResponse(STATIC_DIR / "receivables.html")


@router.get("/api/receivables/meta")
def meta(
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    _ensure_db_ledger_ready(db)
    regions = [
        r[0]
        for r in db.query(models.LicenseHolder.region)
        .filter(models.LicenseHolder.region.isnot(None), models.LicenseHolder.region != "")
        .distinct()
        .order_by(models.LicenseHolder.region.asc())
        .all()
    ]
    return {"regions": regions, "account_types": list(ACCOUNT_FEES), "contact_statuses": sorted(CONTACT_STATUSES), "ledger_mode": LEDGER_MODE, "excel_reupload_required": False, "baseline_key": LEGACY_BASELINE_KEY}


@router.get("/api/receivables/sync")
def sync_receivables(
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    result = _sync_all(db)
    return {"ok": True, **result}


@router.get("/api/receivables/verify")
def verify_legacy_import(
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    _ensure_db_ledger_ready(db)
    payload = json.loads(DATA_FILE.read_text(encoding="utf-8")) if DATA_FILE.exists() else {}
    _load_seed()
    seed_rows = _seed_cache or []

    def signed_stats(values):
        vals = [int(v or 0) for v in values]
        return {
            "arrears_members": sum(1 for v in vals if v > 0),
            "arrears_total": sum(v for v in vals if v > 0),
            "prepaid_members": sum(1 for v in vals if v < 0),
            "prepaid_total": sum(-v for v in vals if v < 0),
            "settled_members": sum(1 for v in vals if v == 0),
            "net_balance": sum(vals),
        }

    seed_stats = signed_stats([r.get("current_arrears") for r in seed_rows])
    seed_by_account = {}
    for r in seed_rows:
        k = r.get("account_type") or "미상"
        seed_by_account[k] = seed_by_account.get(k, 0) + 1

    profiles = db.query(ReceivableProfile).filter(ReceivableProfile.legacy_source_row.isnot(None)).all()
    db_stats = signed_stats([p.legacy_balance for p in profiles])
    db_by_account = {}
    for p in profiles:
        db_by_account[p.account_type] = db_by_account.get(p.account_type, 0) + 1

    charges_sq, payments_sq = _charge_payment_subqueries(db)
    balance_expr = (
        func.coalesce(ReceivableProfile.legacy_balance, 0)
        + func.coalesce(charges_sq.c.charge_total, 0)
        - func.coalesce(payments_sq.c.payment_total, 0)
    )
    current_balances = [
        int(v or 0) for (v,) in (
            db.query(balance_expr)
            .outerjoin(charges_sq, charges_sq.c.member_id == ReceivableProfile.member_id)
            .outerjoin(payments_sq, payments_sq.c.member_id == ReceivableProfile.member_id)
            .filter(ReceivableProfile.legacy_source_row.isnot(None))
            .all()
        )
    ]

    marker = _baseline_marker(db)
    marker_value = {}
    if marker and marker.value:
        try:
            marker_value = json.loads(marker.value)
        except Exception:
            marker_value = {"raw": marker.value}

    return {
        "system": {
            "ledger_mode": LEDGER_MODE,
            "excel_reupload_required": False,
            "baseline_applied": bool(marker),
            "baseline": marker_value,
        },
        "legacy_json": {
            "source_filename": payload.get("source_filename", ""),
            "source_sha256": payload.get("source_sha256", ""),
            "row_count": len(seed_rows),
            **seed_stats,
            "by_account_type": seed_by_account,
            "excluded_rows": payload.get("excluded_rows", []),
        },
        "current_db_legacy_snapshot": {
            "matched_profile_count": len(profiles),
            **db_stats,
            "by_account_type": db_by_account,
        },
        "current_db_after_program_activity": signed_stats(current_balances),
        "match": {
            "row_count_matches": len(seed_rows) == len(profiles),
            "legacy_net_balance_matches": seed_stats["net_balance"] == db_stats["net_balance"],
            "legacy_arrears_total_matches": seed_stats["arrears_total"] == db_stats["arrears_total"],
            "legacy_prepaid_total_matches": seed_stats["prepaid_total"] == db_stats["prepaid_total"],
        },
    }


@router.get("/api/receivables/summary")
def summary(
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    _ensure_db_ledger_ready(db)
    _ensure_current_month_billing(db)
    charges_sq, payments_sq = _charge_payment_subqueries(db)
    balance_expr = (
        func.coalesce(ReceivableProfile.legacy_balance, 0)
        + func.coalesce(charges_sq.c.charge_total, 0)
        - func.coalesce(payments_sq.c.payment_total, 0)
    )
    positive_balance = case((balance_expr > 0, balance_expr), else_=0)
    prepaid_balance = case((balance_expr < 0, -balance_expr), else_=0)
    current_closure_sq = _current_closure_subquery(db)
    active_cond = _receivable_active_sql(current_closure_sq)
    closed_cond = _receivable_closed_sql(current_closure_sq)
    today_iso = datetime.now(KST).date().isoformat()

    row = (
        db.query(
            func.coalesce(func.sum(case((active_cond, 1), else_=0)), 0).label("active_members"),
            func.coalesce(func.sum(case((and_(active_cond, balance_expr > 0), 1), else_=0)), 0).label("active_arrears_members"),
            func.coalesce(func.sum(case((active_cond, positive_balance), else_=0)), 0).label("active_arrears_total"),
            func.coalesce(func.sum(case((and_(active_cond, balance_expr < 0), 1), else_=0)), 0).label("active_prepaid_members"),
            func.coalesce(func.sum(case((active_cond, prepaid_balance), else_=0)), 0).label("active_prepaid_total"),
            func.coalesce(func.sum(case((closed_cond, 1), else_=0)), 0).label("closed_members"),
            func.coalesce(func.sum(case((and_(closed_cond, balance_expr > 0), 1), else_=0)), 0).label("closed_arrears_members"),
            func.coalesce(func.sum(case((closed_cond, positive_balance), else_=0)), 0).label("closed_arrears_total"),
            func.coalesce(func.sum(case((and_(closed_cond, balance_expr < 0), 1), else_=0)), 0).label("closed_prepaid_members"),
            func.coalesce(func.sum(case((closed_cond, prepaid_balance), else_=0)), 0).label("closed_prepaid_total"),
        )
        .select_from(ReceivableProfile)
        .join(models.LicenseHolder, models.LicenseHolder.id == ReceivableProfile.member_id)
        .outerjoin(charges_sq, charges_sq.c.member_id == ReceivableProfile.member_id)
        .outerjoin(payments_sq, payments_sq.c.member_id == ReceivableProfile.member_id)
        .outerjoin(current_closure_sq, current_closure_sq.c.closure_id == models.LicenseHolder.closure_id)
        .one()
    )

    # 오늘 수납 KPI도 현재잔액과 동일한 컷오버 규칙을 사용한다.
    # 2026-08-31 최신 snapshot 확정 전에 이미 DB에 있던 legacy 수납은
    # snapshot에 포함된 과거 수납이므로 오늘 수납으로 다시 세지 않는다.
    today_paid = (
        db.query(func.coalesce(func.sum(ReceivablePayment.amount), 0))
        .join(ReceivableProfile, ReceivableProfile.member_id == ReceivablePayment.member_id)
        .filter(
            ReceivablePayment.payment_date == today_iso,
            ReceivablePayment.cancelled_at.is_(None),
            or_(ReceivablePayment.method.is_(None), ReceivablePayment.method != "잔액수정"),
            _legacy_effective_payment_sql_condition(),
        )
        .scalar()
        or 0
    )
    # 부과대기는 legacy 기존회원이 아니라 실제 신규등록자만 집계한다.
    pending_rows = (
        db.query(ReceivableProfile, models.LicenseHolder)
        .join(models.LicenseHolder, models.LicenseHolder.id == ReceivableProfile.member_id)
        .outerjoin(current_closure_sq, current_closure_sq.c.closure_id == models.LicenseHolder.closure_id)
        .filter(ReceivableProfile.legacy_source_row.is_(None))
        .filter(ReceivableProfile.first_charge_date > today_iso)
        .filter(_receivable_active_sql(current_closure_sq))
        .all()
    )
    pending_members = sum(1 for p, m in pending_rows if _is_true_new_member(m, p))

    _schedule_background_sync(background_tasks)
    return {
        "active_members": int(row.active_members or 0),
        "active_arrears_members": int(row.active_arrears_members or 0),
        "active_arrears_total": int(row.active_arrears_total or 0),
        "active_prepaid_members": int(row.active_prepaid_members or 0),
        "active_prepaid_total": int(row.active_prepaid_total or 0),
        "closed_members": int(row.closed_members or 0),
        "closed_arrears_members": int(row.closed_arrears_members or 0),
        "closed_arrears_total": int(row.closed_arrears_total or 0),
        "closed_prepaid_members": int(row.closed_prepaid_members or 0),
        "closed_prepaid_total": int(row.closed_prepaid_total or 0),
        "pending_members": int(pending_members),
        "today_paid": int(today_paid),
    }




@router.get("/api/receivables/dashboard")
def receivables_dashboard(
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """미수금현황 대시보드용 읽기 전용 집계.

    기존 원장/부과/수납 데이터를 변경하지 않고 현재 잔액, 지역/계정 분포,
    최신 연락기록, 고액 미수 순위를 한 번에 반환한다.
    """
    _ensure_db_ledger_ready(db)
    _ensure_current_month_billing(db)
    charges_sq, payments_sq = _charge_payment_subqueries(db)
    latest_contact_sq = _latest_contact_subquery(db)
    balance_expr = (
        func.coalesce(ReceivableProfile.legacy_balance, 0)
        + func.coalesce(charges_sq.c.charge_total, 0)
        - func.coalesce(payments_sq.c.payment_total, 0)
    ).label("balance")

    rows = (
        db.query(
            models.LicenseHolder.id.label("member_id"),
            models.LicenseHolder.name.label("name"),
            models.LicenseHolder.region.label("region"),
            models.LicenseHolder.vehicle_number.label("vehicle_number"),
            models.LicenseHolder.status.label("member_status"),
            models.LicenseHolder.deleted_at.label("deleted_at"),
            ReceivableProfile.account_type.label("account_type"),
            balance_expr,
            latest_contact_sq.c.status.label("contact_status"),
            latest_contact_sq.c.contact_date.label("last_contact_date"),
        )
        .join(ReceivableProfile, ReceivableProfile.member_id == models.LicenseHolder.id)
        .outerjoin(charges_sq, charges_sq.c.member_id == ReceivableProfile.member_id)
        .outerjoin(payments_sq, payments_sq.c.member_id == ReceivableProfile.member_id)
        .outerjoin(latest_contact_sq, latest_contact_sq.c.member_id == models.LicenseHolder.id)
        .filter(models.LicenseHolder.deleted_at.is_(None))
        .all()
    )

    active_arrears = []
    active_prepaid_total = 0
    active_prepaid_members = 0
    closed_arrears_total = 0
    closed_arrears_members = 0
    region_map = {}
    account_map = {}
    arrears_contact_status = {k: 0 for k in ["연락완료", "문자발송", "재연락 필요", "부재", "미연락"]}
    contact_overview_status = {k: 0 for k in ["연락완료", "문자발송", "재연락 필요", "부재", "미연락"]}
    total_contacted = 0
    recent_7_days = 0
    recent_cutoff = (datetime.now(KST).date() - timedelta(days=6)).isoformat()

    for r in rows:
        bal = int(r.balance or 0)
        is_active = r.member_status in (None, "active")
        is_closed = r.member_status == "closed"
        contact_exists = bool(r.last_contact_date)
        contact_status = (r.contact_status or "미연락").strip() or "미연락"

        if contact_exists:
            total_contacted += 1
            contact_overview_status[contact_status] = contact_overview_status.get(contact_status, 0) + 1
            if str(r.last_contact_date) >= recent_cutoff:
                recent_7_days += 1

        if is_closed and bal > 0:
            closed_arrears_total += bal
            closed_arrears_members += 1
        if is_active and bal < 0:
            active_prepaid_total += -bal
            active_prepaid_members += 1
        if not (is_active and bal > 0):
            continue

        item = {
            "member_id": int(r.member_id),
            "name": r.name or "",
            "region": r.region or "미지정",
            "vehicle_number": r.vehicle_number or "",
            "account_type": r.account_type or "기타",
            "balance": bal,
            "contact_status": contact_status,
            "last_contact_date": r.last_contact_date or "",
            "contacted": contact_exists,
        }
        active_arrears.append(item)
        arrears_contact_status[contact_status] = arrears_contact_status.get(contact_status, 0) + 1

        rg = region_map.setdefault(item["region"], {"region": item["region"], "arrears_total": 0, "arrears_members": 0, "contacted_members": 0})
        rg["arrears_total"] += bal
        rg["arrears_members"] += 1
        rg["contacted_members"] += 1 if contact_exists else 0

        ac = account_map.setdefault(item["account_type"], {"account_type": item["account_type"], "arrears_total": 0, "arrears_members": 0})
        ac["arrears_total"] += bal
        ac["arrears_members"] += 1

    active_total = sum(x["balance"] for x in active_arrears)
    active_members = len(active_arrears)
    contacted_members = sum(1 for x in active_arrears if x["contacted"])
    over_100 = [x for x in active_arrears if x["balance"] >= 100000]

    regions = []
    for x in region_map.values():
        members = int(x["arrears_members"] or 0)
        x["average_arrears"] = int(round(x["arrears_total"] / members)) if members else 0
        x["contact_rate"] = round((x["contacted_members"] / members * 100), 1) if members else 0
        x["share_pct"] = round((x["arrears_total"] / active_total * 100), 1) if active_total else 0
        regions.append(x)
    regions.sort(key=lambda x: (-x["arrears_total"], x["region"]))

    account_order = {"협회비": 0, "관리비": 1, "70세": 2}
    accounts = []
    for x in account_map.values():
        members = int(x["arrears_members"] or 0)
        x["average_arrears"] = int(round(x["arrears_total"] / members)) if members else 0
        x["share_pct"] = round((x["arrears_total"] / active_total * 100), 1) if active_total else 0
        accounts.append(x)
    accounts.sort(key=lambda x: (account_order.get(x["account_type"], 99), -x["arrears_total"]))

    band_defs = [
        ("3만원 미만", 0, 30000),
        ("3만~10만원", 30000, 100000),
        ("10만~30만원", 100000, 300000),
        ("30만~100만원", 300000, 1000000),
        ("100만원 이상", 1000000, None),
    ]
    balance_bands = []
    for label, lo, hi in band_defs:
        vals = [x["balance"] for x in active_arrears if x["balance"] >= lo and (hi is None or x["balance"] < hi)]
        balance_bands.append({"label": label, "members": len(vals), "amount": int(sum(vals))})

    top_arrears = sorted(active_arrears, key=lambda x: (-x["balance"], x["name"]))[:10]

    return {
        "summary": {
            "active_arrears_total": int(active_total),
            "active_arrears_members": int(active_members),
            "average_arrears": int(round(active_total / active_members)) if active_members else 0,
            "contacted_members": int(contacted_members),
            "contact_rate": round((contacted_members / active_members * 100), 1) if active_members else 0,
            "over_100k_members": len(over_100),
            "over_100k_total": int(sum(x["balance"] for x in over_100)),
            "closed_arrears_total": int(closed_arrears_total),
            "closed_arrears_members": int(closed_arrears_members),
            "active_prepaid_total": int(active_prepaid_total),
            "active_prepaid_members": int(active_prepaid_members),
        },
        "regions": regions,
        "accounts": accounts,
        "balance_bands": balance_bands,
        "arrears_contact_status": arrears_contact_status,
        "top_arrears": top_arrears,
        "contact_overview": {
            "total_contacted": int(total_contacted),
            "recent_7_days": int(recent_7_days),
            "status_counts": contact_overview_status,
        },
    }

def _analysis_month_end(month_key: str) -> date:
    y, m = [int(x) for x in month_key.split("-", 1)]
    return date(y, m, calendar.monthrange(y, m)[1])


def _analysis_prev_month(month_key: str) -> str:
    y, m = [int(x) for x in month_key.split("-", 1)]
    if m == 1:
        return f"{y - 1:04d}-12"
    return f"{y:04d}-{m - 1:02d}"


def _analysis_shift_month(month_key: str, delta: int) -> str:
    y, m = [int(x) for x in month_key.split("-", 1)]
    idx = y * 12 + (m - 1) + delta
    return f"{idx // 12:04d}-{idx % 12 + 1:02d}"


def _legacy_month_row(profile: ReceivableProfile, month_key: str):
    y, m = [int(x) for x in month_key.split("-", 1)]
    if y != LEGACY_YEAR:
        return None
    for row in (profile.legacy_months or []):
        try:
            if int(row.get("month") or 0) == m:
                return row
        except Exception:
            continue
    return None


@router.get("/api/receivables/monthly-analysis")
def monthly_analysis(
    from_month: Optional[str] = Query(None),
    to_month: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """월별 미수금 증감·부과·수납 성과를 비교한다.

    사용자가 원하는 핵심은 '어느 월 대비 미수금이 얼마나 줄거나 늘었는지'와
    '그 기간에 얼마를 부과했고 얼마를 실제로 수납했는지'다. 따라서 연락률·지역
    랭킹이 아니라 월말 원장 스냅샷과 월별 현금흐름을 중심으로 읽기 전용 집계를
    반환한다. 기존 수납/부과/폐업 저장 로직은 변경하지 않는다.
    """
    _ensure_db_ledger_ready(db)
    _ensure_current_month_billing(db)

    today_d = datetime.now(KST).date()
    current_key = f"{today_d.year:04d}-{today_d.month:02d}"
    available_start = f"{LEGACY_YEAR:04d}-01"

    def normalize_month(value: Optional[str], field_name: str) -> Optional[str]:
        if value is None or str(value).strip() == "":
            return None
        value = str(value).strip()
        if not re.fullmatch(r"\d{4}-\d{2}", value):
            raise HTTPException(status_code=400, detail=f"{field_name} 형식은 YYYY-MM 이어야 합니다.")
        yy, mm = [int(x) for x in value.split("-", 1)]
        if not (1 <= mm <= 12):
            raise HTTPException(status_code=400, detail=f"{field_name} 월 값이 올바르지 않습니다.")
        return f"{yy:04d}-{mm:02d}"

    selected_to = normalize_month(to_month, "비교월") or current_key
    if selected_to < available_start or selected_to > current_key:
        raise HTTPException(status_code=400, detail=f"비교월은 {available_start}~{current_key} 범위에서 선택하세요.")

    default_from = _analysis_prev_month(selected_to)
    if default_from < available_start:
        default_from = available_start
    selected_from = normalize_month(from_month, "기준월") or default_from
    if selected_from < available_start or selected_from > current_key:
        raise HTTPException(status_code=400, detail=f"기준월은 {available_start}~{current_key} 범위에서 선택하세요.")
    if selected_from > selected_to:
        raise HTTPException(status_code=400, detail="기준월은 비교월보다 앞선 월이어야 합니다.")

    profiles = db.query(ReceivableProfile).all()
    profile_by_member = {int(p.member_id): p for p in profiles}
    member_ids = set(profile_by_member)

    # 프로그램에서 실제로 추가된 월 부과. legacy 1~8월은 기준원장에 이미 포함되어 제외한다.
    charges_by_member = {}
    program_charge_by_month = {}
    if member_ids:
        for c in db.query(ReceivableCharge).filter(ReceivableCharge.member_id.in_(member_ids)).all():
            p = profile_by_member.get(int(c.member_id))
            if not p:
                continue
            billing_month = str(c.billing_month or "")
            if p.legacy_source_row is not None and billing_month <= LEGACY_DATA_THROUGH_KEY:
                continue
            amount = int(c.amount or 0)
            charges_by_member.setdefault(int(c.member_id), []).append((billing_month, amount))
            program_charge_by_month[billing_month] = program_charge_by_month.get(billing_month, 0) + amount

    # 프로그램 수납 및 잔액수정. 실제 수납과 금액수정은 분리한다.
    payments_by_member = {}
    program_paid_by_month = {}
    balance_adjustment_by_month = {}
    if member_ids:
        payment_rows = (
            db.query(ReceivablePayment)
            .filter(ReceivablePayment.member_id.in_(member_ids), ReceivablePayment.cancelled_at.is_(None))
            .all()
        )
        for pay in payment_rows:
            profile = profile_by_member.get(int(pay.member_id))
            if profile is not None and _legacy_payment_is_snapshot_covered(profile, pay):
                continue
            pd = str(pay.payment_date or "")
            if len(pd) < 7:
                continue
            mk = pd[:7]
            amount = int(pay.amount or 0)
            method = str(pay.method or "")
            payments_by_member.setdefault(int(pay.member_id), []).append((pd, amount, method))
            if method == "잔액수정":
                # 잔액 = ... - payment.amount 이므로 실제 잔액 변화량은 -amount.
                balance_adjustment_by_month[mk] = balance_adjustment_by_month.get(mk, 0) - amount
            else:
                program_paid_by_month[mk] = program_paid_by_month.get(mk, 0) + amount

    snapshot_cache = {}

    def snapshot(profile: ReceivableProfile, month_key: str) -> int:
        cache_key = (int(profile.member_id), month_key)
        if cache_key in snapshot_cache:
            return snapshot_cache[cache_key]

        end_iso = _analysis_month_end(month_key).isoformat()
        y, m = [int(x) for x in month_key.split("-", 1)]

        if profile.legacy_source_row is not None:
            if (y, m) == (LEGACY_YEAR - 1, 12):
                seed = _seed_for_profile(profile) or {}
                base = int(seed.get("carryover") or 0)
            elif y == LEGACY_YEAR and 1 <= m <= LEGACY_DATA_THROUGH_MONTH:
                reconstructed = _legacy_reconstructed_balance(profile, m)
                base = int(profile.legacy_balance or 0) if reconstructed is None else int(reconstructed)
            elif (y, m) < (LEGACY_YEAR, 1):
                base = 0
            else:
                base = int(profile.legacy_balance or 0)
        else:
            base = 0

        for billing_month, amount in charges_by_member.get(int(profile.member_id), []):
            if billing_month and billing_month <= month_key:
                base += int(amount or 0)
        for payment_date, amount, _method in payments_by_member.get(int(profile.member_id), []):
            if payment_date and payment_date <= end_iso:
                base -= int(amount or 0)

        snapshot_cache[cache_key] = int(base)
        return int(base)

    aggregate_cache = {}

    def aggregate(month_key: str):
        if month_key in aggregate_cache:
            return aggregate_cache[month_key]
        arrears_total = 0
        arrears_members = 0
        prepaid_total = 0
        prepaid_members = 0
        balances = {}
        for p in profiles:
            b = snapshot(p, month_key)
            balances[int(p.member_id)] = b
            if b > 0:
                arrears_total += b
                arrears_members += 1
            elif b < 0:
                prepaid_total += -b
                prepaid_members += 1
        out = {
            "arrears_total": int(arrears_total),
            "arrears_members": int(arrears_members),
            "prepaid_total": int(prepaid_total),
            "prepaid_members": int(prepaid_members),
            "balances": balances,
        }
        aggregate_cache[month_key] = out
        return out

    def movement_between(start_data, end_data):
        increased_amount = 0
        decreased_amount = 0
        increased_members = 0
        decreased_members = 0
        new_arrears_members = 0
        settled_members = 0
        for mid in member_ids:
            pb = max(int(start_data["balances"].get(mid, 0)), 0)
            cb = max(int(end_data["balances"].get(mid, 0)), 0)
            diff = cb - pb
            if diff > 0:
                increased_amount += diff
                increased_members += 1
            elif diff < 0:
                decreased_amount += -diff
                decreased_members += 1
            if pb <= 0 < cb:
                new_arrears_members += 1
            if pb > 0 and cb <= 0:
                settled_members += 1
        return {
            "increased_amount": int(increased_amount),
            "decreased_amount": int(decreased_amount),
            "increased_members": int(increased_members),
            "decreased_members": int(decreased_members),
            "new_arrears_members": int(new_arrears_members),
            "settled_members": int(settled_members),
        }

    legacy_flow_cache = {}

    def month_flow(month_key: str):
        if month_key in legacy_flow_cache:
            return legacy_flow_cache[month_key]
        yy, mm = [int(x) for x in month_key.split("-", 1)]
        legacy_paid = 0
        legacy_charge = 0
        if yy == LEGACY_YEAR and mm <= LEGACY_DATA_THROUGH_MONTH:
            for p in profiles:
                if p.legacy_source_row is None:
                    continue
                lr = _legacy_month_row(p, month_key)
                if not lr:
                    continue
                if lr.get("payment") is not None:
                    legacy_paid += int(lr.get("payment") or 0)
                monthly_charge = lr.get("monthly_charge")
                if monthly_charge is None and lr.get("billed_total") is not None:
                    monthly_charge = ACCOUNT_FEES.get(p.account_type)
                if monthly_charge is not None:
                    legacy_charge += int(monthly_charge or 0)
        out = {
            "payments": int(legacy_paid + program_paid_by_month.get(month_key, 0)),
            "charges": int(legacy_charge + program_charge_by_month.get(month_key, 0)),
            "balance_adjustment": int(balance_adjustment_by_month.get(month_key, 0)),
        }
        legacy_flow_cache[month_key] = out
        return out

    # 전체 월별 성과 이력. 월말 미수, 전월대비 증감, 부과·수납, 수납/부과율을 함께 준다.
    history = []
    mk = available_start
    while mk <= current_key:
        cur = aggregate(mk)
        prev_key = _analysis_prev_month(mk)
        prev = aggregate(prev_key)
        mv = movement_between(prev, cur)
        flow = month_flow(mk)
        net_change = int(cur["arrears_total"] - prev["arrears_total"])
        net_change_pct = round(net_change / prev["arrears_total"] * 100, 1) if prev["arrears_total"] else None
        collection_ratio = round(flow["payments"] / flow["charges"] * 100, 1) if flow["charges"] else None
        net_reduction_rate = round((prev["arrears_total"] - cur["arrears_total"]) / prev["arrears_total"] * 100, 1) if prev["arrears_total"] else None
        history.append({
            "month": mk,
            "label": f"{int(mk[:4])}년 {int(mk[5:7])}월",
            "start_arrears": int(prev["arrears_total"]),
            "end_arrears": int(cur["arrears_total"]),
            "arrears_members": int(cur["arrears_members"]),
            "prepaid_total": int(cur["prepaid_total"]),
            "payments": int(flow["payments"]),
            "charges": int(flow["charges"]),
            "balance_adjustment": int(flow["balance_adjustment"]),
            "collection_ratio": collection_ratio,
            "net_change": net_change,
            "net_change_pct": net_change_pct,
            "net_reduction_rate": net_reduction_rate,
            "increased_amount": int(mv["increased_amount"]),
            "decreased_amount": int(mv["decreased_amount"]),
            "increased_members": int(mv["increased_members"]),
            "decreased_members": int(mv["decreased_members"]),
            "new_arrears_members": int(mv["new_arrears_members"]),
            "settled_members": int(mv["settled_members"]),
            "arrears_member_change": int(cur["arrears_members"] - prev["arrears_members"]),
        })
        mk = _analysis_shift_month(mk, 1)

    start_data = aggregate(selected_from)
    end_data = aggregate(selected_to)
    selected_movement = movement_between(start_data, end_data)
    net_change = int(end_data["arrears_total"] - start_data["arrears_total"])
    net_change_pct = round(net_change / start_data["arrears_total"] * 100, 1) if start_data["arrears_total"] else None

    # 월말 A → 월말 B 비교이므로 기간 거래는 A 다음 달부터 B월까지 합산한다.
    period_keys = []
    if selected_from < selected_to:
        pk = _analysis_shift_month(selected_from, 1)
        while pk <= selected_to:
            period_keys.append(pk)
            pk = _analysis_shift_month(pk, 1)
    else:
        period_keys = [selected_to]

    period_payments = sum(month_flow(x)["payments"] for x in period_keys)
    period_charges = sum(month_flow(x)["charges"] for x in period_keys)
    period_adjustment = sum(month_flow(x)["balance_adjustment"] for x in period_keys)
    payment_charge_ratio = round(period_payments / period_charges * 100, 1) if period_charges else None
    gross_recovery_rate = round(selected_movement["decreased_amount"] / start_data["arrears_total"] * 100, 1) if start_data["arrears_total"] else None
    settled_rate = round(selected_movement["settled_members"] / start_data["arrears_members"] * 100, 1) if start_data["arrears_members"] else None

    # 인사이트: 전월 대비 미수금이 가장 많이 줄고/늘어난 달, 수납/부과율 최고 월.
    comparable = [x for x in history if x["month"] > available_start or x["start_arrears"] > 0]
    reduction_rows = [x for x in comparable if int(x.get("net_change") or 0) < 0]
    increase_rows = [x for x in comparable if int(x.get("net_change") or 0) > 0]
    best_reduction = min(reduction_rows, key=lambda x: x["net_change"], default=None)
    worst_increase = max(increase_rows, key=lambda x: x["net_change"], default=None)
    ratio_rows = [x for x in history if x["collection_ratio"] is not None]
    best_collection = max(ratio_rows, key=lambda x: x["collection_ratio"], default=None)

    available_months = [{"value": x["month"], "label": x["label"]} for x in history]

    return {
        "basis": "전체 원장(활성+폐업) · 월말 기준",
        "available_months": available_months,
        "period": {"from": selected_from, "to": selected_to, "months": len(period_keys)},
        "from": {
            "month": selected_from,
            "arrears_total": int(start_data["arrears_total"]),
            "arrears_members": int(start_data["arrears_members"]),
            "average_arrears": int(round(start_data["arrears_total"] / start_data["arrears_members"])) if start_data["arrears_members"] else 0,
            "prepaid_total": int(start_data["prepaid_total"]),
        },
        "to": {
            "month": selected_to,
            "arrears_total": int(end_data["arrears_total"]),
            "arrears_members": int(end_data["arrears_members"]),
            "average_arrears": int(round(end_data["arrears_total"] / end_data["arrears_members"])) if end_data["arrears_members"] else 0,
            "prepaid_total": int(end_data["prepaid_total"]),
        },
        "comparison": {
            **selected_movement,
            "net_change": net_change,
            "net_change_pct": net_change_pct,
            "arrears_member_change": int(end_data["arrears_members"] - start_data["arrears_members"]),
            "period_payments": int(period_payments),
            "period_charges": int(period_charges),
            "period_adjustment": int(period_adjustment),
            "payment_charge_ratio": payment_charge_ratio,
            "gross_recovery_rate": gross_recovery_rate,
            "settled_rate": settled_rate,
        },
        "history": history,
        "insights": {
            "best_reduction": best_reduction,
            "worst_increase": worst_increase,
            "best_collection": best_collection,
        },
        # 구버전 프론트 호환 필드. 새 화면에서는 comparison/history를 사용한다.
        "previous": {"arrears_total": int(start_data["arrears_total"]), "arrears_members": int(start_data["arrears_members"])},
        "current": {"arrears_total": int(end_data["arrears_total"]), "arrears_members": int(end_data["arrears_members"]), "prepaid_total": int(end_data["prepaid_total"])},
        "movement": {
            "payments": int(period_payments),
            "charges": int(period_charges),
            "balance_adjustment": int(period_adjustment),
            **selected_movement,
            "net_change": net_change,
            "net_change_pct": net_change_pct,
            "arrears_member_change": int(end_data["arrears_members"] - start_data["arrears_members"]),
        },
        "trend": [{"month": x["month"], "label": f"{int(x['month'][5:7])}월", "arrears_total": x["end_arrears"], "arrears_members": x["arrears_members"]} for x in history],
    }

def _normalize_closure_type(v: str) -> str:
    v = str(v or "").strip()
    return "폐업" if v == "폐지" else (v or "폐업")


def _hydrate_closure_records(db: Session, closure_rows):
    """폐업관리 화면용. 폐업현황(closures) 행 자체를 기준으로 표시하고,
    수납 원장 연결은 member_id 우선 + 폐업현황 라우터의 안전 매칭 규칙으로 보강한다.

    중요: 폐업현황 행이 원장과 연결되지 않아도 행 자체는 숨기지 않는다.
    관리번호/구분/접수일/처리일/양수인/이관지역은 Closure가 원본이다.
    """
    if not closure_rows:
        return []

    # 기존 폐업현황 화면과 동일한 안전 매칭 규칙을 재사용한다.
    from app.routers.closures import _build_member_lookup, _norm_vn

    by_id, by_vehicle, by_resident = _build_member_lookup(db, closure_rows)

    # 현재 폐업건은 LicenseHolder.closure_id가 가리키는 관계가 가장 신뢰할 수 있는 연결이다.
    # 성명/차량번호 매칭보다 먼저 사용해서 과거 동명이인/차량 재사용 오연결을 막는다.
    closure_ids = [int(c.id) for c in closure_rows if getattr(c, "id", None) is not None]
    current_member_by_closure = {}
    if closure_ids:
        for m in (
            db.query(models.LicenseHolder)
            .filter(models.LicenseHolder.closure_id.in_(closure_ids))
            .all()
        ):
            if getattr(m, "closure_id", None) is not None:
                current_member_by_closure[int(m.closure_id)] = m

    def _strict_closure_member(c):
        # 1) Closure.member_id 직접연결
        mid = getattr(c, "member_id", None)
        if mid and mid in by_id:
            return by_id[mid]

        # 2) 주민번호 완전일치. 중복이면 성명까지 완전일치해야 한다.
        rn = (getattr(c, "resident_number", "") or "").strip()
        if rn and rn in by_resident:
            candidates = by_resident[rn]
            if len(candidates) == 1:
                return candidates[0]
            name = (getattr(c, "name", "") or "").strip()
            exact = [m for m in candidates if name and (m.name or "").strip() == name]
            if len(exact) == 1:
                return exact[0]

        # 3) 차량번호는 반드시 성명까지 동시에 완전일치해야 한다.
        #    차량번호만 유일하다는 이유로 연결하지 않는다.
        #    예: 폐업 이종환 85바5040 ↔ 현역 이종한 85배5040 같은 오연결 차단.
        vn = _norm_vn(getattr(c, "vehicle_number", "") or "")
        name = (getattr(c, "name", "") or "").strip()
        if vn and name and vn in by_vehicle:
            exact = [m for m in by_vehicle[vn] if (m.name or "").strip() == name]
            if len(exact) == 1:
                return exact[0]
        return None

    linked = {}
    member_ids = set()
    for c in closure_rows:
        m = current_member_by_closure.get(int(c.id)) if getattr(c, "id", None) is not None else None
        if m is None:
            m = _strict_closure_member(c)
        linked[c.id] = m
        if m is not None:
            member_ids.add(m.id)

    profile_map = {}
    balance_map = {}
    contact_map = {}
    if member_ids:
        charges_sq, payments_sq = _charge_payment_subqueries(db)
        balance_expr = (
            func.coalesce(ReceivableProfile.legacy_balance, 0)
            + func.coalesce(charges_sq.c.charge_total, 0)
            - func.coalesce(payments_sq.c.payment_total, 0)
        ).label("balance")
        for profile, balance in (
            db.query(ReceivableProfile, balance_expr)
            .outerjoin(charges_sq, charges_sq.c.member_id == ReceivableProfile.member_id)
            .outerjoin(payments_sq, payments_sq.c.member_id == ReceivableProfile.member_id)
            .filter(ReceivableProfile.member_id.in_(member_ids))
            .all()
        ):
            profile_map[profile.member_id] = profile
            balance_map[profile.member_id] = int(balance or 0)

        latest_contact_sq = _latest_contact_subquery(db)
        for mid, status, contact_date in (
            db.query(
                latest_contact_sq.c.member_id,
                latest_contact_sq.c.status,
                latest_contact_sq.c.contact_date,
            )
            .filter(latest_contact_sq.c.member_id.in_(member_ids))
            .all()
        ):
            contact_map[int(mid)] = (status or "미연락", contact_date or "")

    items = []
    for c in closure_rows:
        m = linked.get(c.id)
        ct = _normalize_closure_type(c.closure_type)
        if m is not None:
            profile = profile_map.get(m.id)
            contact_status, last_contact_date = contact_map.get(m.id, ("미연락", ""))
            if profile is not None:
                display_balance = int(balance_map.get(m.id, 0) or 0)
                close_d = _parse_date(getattr(c, "closure_date", None))
                if close_d is not None:
                    display_balance += _legacy_balance_as_of(profile, close_d) - int(profile.legacy_balance or 0)
                item = _serialize_member(
                    m, profile, display_balance,
                    contact_status, last_contact_date,
                    c.id, c.management_number or "", c.closure_date or "", ct,
                    c.reason or "", c.transferee or "", c.transfer_region or "",
                    c.receipt_date or "",
                )
            else:
                item = {
                    "member_id": m.id,
                    "name": m.name or c.name or "",
                    "vehicle_number": m.vehicle_number or c.vehicle_number or "",
                    "region": m.region or c.region or "",
                    "mobile": getattr(c, "mobile", None) or getattr(m, "mobile", None) or "",
                    "phone": getattr(c, "phone", None) or getattr(m, "phone", None) or "",
                    "address": getattr(m, "address", None) or "",
                    "category": m.category or "",
                    "account_type": "협회비" if is_association_member(getattr(m, "membership_date", None)) else "관리비",
                    "unit_fee": 0, "vehicle_count": 1,
                    "membership_status": "가입" if is_association_member(getattr(m, "membership_date", None)) else "미가입",
                    "membership_date": m.membership_date or "",
                    "member_status": ct, "active": False,
                    "balance": None, "arrears_amount": 0, "prepaid_amount": 0,
                    "first_charge_date": "", "billing_state": "폐업기록",
                    "legacy_member": False,
                    "contact_status": contact_status, "last_contact_date": last_contact_date,
                }
            # 폐업관리 목록의 식별정보는 현재 회원정보가 아니라 해당 폐업현황 행을 우선한다.
            item.update({
                "name": c.name or item.get("name", ""),
                "vehicle_number": c.vehicle_number or item.get("vehicle_number", ""),
                "region": c.region or item.get("region", ""),
                "mobile": getattr(c, "mobile", None) or item.get("mobile", ""),
                "phone": getattr(c, "phone", None) or item.get("phone", ""),
            })
        else:
            # 과거 폐업현황은 회원마스터/미수원장과 연결되지 않아도 반드시 보여준다.
            item = {
                "member_id": None,
                "name": c.name or "",
                "vehicle_number": c.vehicle_number or "",
                "region": c.region or "",
                "mobile": getattr(c, "mobile", None) or "",
                "phone": getattr(c, "phone", None) or "",
                "address": "",
                "category": "", "account_type": "", "unit_fee": 0, "vehicle_count": 1,
                "membership_status": "", "membership_date": "",
                "member_status": ct, "active": False,
                "balance": None, "arrears_amount": 0, "prepaid_amount": 0,
                "first_charge_date": "", "billing_state": "폐업기록",
                "legacy_member": False,
                "contact_status": "", "last_contact_date": "",
            }

        item.update({
            "closure_id": c.id,
            "closure_management_number": c.management_number or "",
            "closure_date": c.closure_date or "",
            "closure_type": ct,
            "closure_reason": c.reason or "",
            "transferee": c.transferee or "",
            "transfer_region": c.transfer_region or "",
            "closure_receipt_date": c.receipt_date or "",
            "closure_record_only": m is None,
            "has_receivable_link": bool(m is not None and profile_map.get(m.id) is not None),
        })
        items.append(item)
    return items


def _list_closure_records(
    db: Session, *, q: str = "", region: str = "", account_type: str = "",
    contact_status: str = "", billing_status: str = "", arrears_only: bool = False,
    closure_mode: str = "current", page: int = 1, limit: int = 50,
):
    """폐업관리 조회.

    current: 현재 회원마스터에서 실제 status='closed'이고 closure_id가 가리키는 현재 폐업건만.
             과거 이력이 있어도 현재 active인 회원은 절대 포함하지 않는다.
    history: 인허가/변경 > 폐업현황(closures) 전체 이력. 과거 7천여 건 원장을 보존 조회한다.
    """
    if closure_mode == "history":
        base = db.query(models.Closure).filter(models.Closure.deleted_at.is_(None))
    else:
        # 현재 폐업관리의 권위 기준은 LicenseHolder.status + LicenseHolder.closure_id.
        # Closure 과거 이력만 존재하는 현재활동 회원은 여기서 완전히 제외된다.
        base = (
            db.query(models.Closure)
            .join(models.LicenseHolder, models.LicenseHolder.closure_id == models.Closure.id)
            .filter(
                models.Closure.deleted_at.is_(None),
                models.LicenseHolder.status == "closed",
                models.LicenseHolder.closure_id.isnot(None),
            )
        )
    base = base.filter(or_(
        and_(models.Closure.vehicle_number.isnot(None), models.Closure.vehicle_number != ""),
        and_(models.Closure.name.isnot(None), models.Closure.name != ""),
    ))
    if region:
        base = base.filter(models.Closure.region == region)
    raw_q = (q or "").strip()
    if raw_q:
        pat = f"%{raw_q}%"
        base = base.filter(or_(
            models.Closure.management_number.ilike(pat),
            models.Closure.name.ilike(pat),
            models.Closure.vehicle_number.ilike(pat),
            models.Closure.region.ilike(pat),
            models.Closure.mobile.ilike(pat),
            models.Closure.phone.ilike(pat),
            models.Closure.transferee.ilike(pat),
            models.Closure.transfer_region.ilike(pat),
            models.Closure.closure_type.ilike(pat),
        ))

    # 폐업현황의 최신 입력 순서가 실제 폐업현황 화면의 최근 자료와 가장 안정적으로 일치한다.
    ordered = base.order_by(models.Closure.id.desc())
    needs_link_filter = bool(account_type or contact_status or billing_status or arrears_only)
    if not needs_link_filter:
        total = base.order_by(None).count()
        closure_rows = ordered.offset((page - 1) * limit).limit(limit).all()
        items = _hydrate_closure_records(db, closure_rows)
    else:
        # 원장/연락 필터를 사용하는 경우에만 연결정보를 보강한 뒤 필터한다.
        # 기본 폐업관리 조회는 위 경로로 DB 페이지네이션하므로 빠르다.
        all_items = _hydrate_closure_records(db, ordered.all())
        def keep(x):
            if account_type and x.get("account_type") != account_type:
                return False
            if contact_status:
                cs = x.get("contact_status") or ""
                if contact_status == "미연락":
                    if cs not in ("", "미연락"):
                        return False
                elif cs != contact_status:
                    return False
            bal = x.get("balance")
            if arrears_only and not (isinstance(bal, (int, float)) and bal > 0):
                return False
            if billing_status == "arrears" and not (isinstance(bal, (int, float)) and bal > 0):
                return False
            if billing_status == "settled" and bal != 0:
                return False
            if billing_status == "prepaid" and not (isinstance(bal, (int, float)) and bal < 0):
                return False
            if billing_status == "pending":
                return False
            return True
        filtered = [x for x in all_items if keep(x)]
        total = len(filtered)
        start = (page - 1) * limit
        items = filtered[start:start + limit]

    return {
        "items": items,
        "count": int(total),
        "page": page,
        "limit": limit,
        "pages": max(1, (int(total) + limit - 1) // limit),
        "source": "closures",
        "closure_mode": closure_mode,
    }


@router.get("/api/receivables/members")
def list_members(
    background_tasks: BackgroundTasks,
    scope: str = Query("active", pattern="^(active|closed|all)$"),
    closure_mode: str = Query("current", pattern="^(current|history)$"),
    arrears_only: bool = False,
    q: str = "",
    region: str = "",
    account_type: str = "",
    contact_status: str = "",
    contacted_only: bool = False,
    billing_status: str = Query("", pattern="^(|pending|arrears|settled|prepaid)$"),
    page: int = Query(1, ge=1),
    limit: int = Query(50, ge=10, le=200),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    # 최신 Excel은 최초 1회만 DB로 이관된다. 이후 신규회원만 빠르게 연결한다.
    _ensure_db_ledger_ready(db)
    created_profiles = _sync_missing_profiles_fast(db, limit=100, allow_legacy_seed=False)
    repaired_profiles = _repair_account_types(db)
    _ensure_current_month_billing(db)
    if created_profiles or repaired_profiles:
        _sync_charges(db)

    # 폐업관리는 반드시 인허가/변경 > 폐업현황(closures) 원장을 직접 기준으로 한다.
    # member_id가 없는 과거 자료도 관리번호/구분과 함께 그대로 보여야 한다.
    if scope == "closed":
        result = _list_closure_records(
            db, q=q, region=region, account_type=account_type,
            contact_status=contact_status, billing_status=billing_status,
            arrears_only=arrears_only, closure_mode=closure_mode, page=page, limit=limit,
        )
        _schedule_background_sync(background_tasks)
        return result

    charges_sq, payments_sq = _charge_payment_subqueries(db)
    latest_contact_sq = _latest_contact_subquery(db)
    current_closure_sq = _current_closure_subquery(db)
    balance_expr = (
        func.coalesce(ReceivableProfile.legacy_balance, 0)
        + func.coalesce(charges_sq.c.charge_total, 0)
        - func.coalesce(payments_sq.c.payment_total, 0)
    ).label("balance")

    query = (
        db.query(
            models.LicenseHolder,
            ReceivableProfile,
            balance_expr,
            latest_contact_sq.c.status.label("contact_status"),
            latest_contact_sq.c.contact_date.label("last_contact_date"),
            current_closure_sq.c.closure_id.label("closure_id"),
            current_closure_sq.c.management_number.label("closure_management_number"),
            current_closure_sq.c.closure_date.label("closure_date"),
            current_closure_sq.c.closure_type.label("closure_type"),
            current_closure_sq.c.reason.label("closure_reason"),
            current_closure_sq.c.transferee.label("transferee"),
            current_closure_sq.c.transfer_region.label("transfer_region"),
            current_closure_sq.c.receipt_date.label("closure_receipt_date"),
        )
        .join(ReceivableProfile, ReceivableProfile.member_id == models.LicenseHolder.id)
        .outerjoin(charges_sq, charges_sq.c.member_id == ReceivableProfile.member_id)
        .outerjoin(payments_sq, payments_sq.c.member_id == ReceivableProfile.member_id)
        .outerjoin(latest_contact_sq, latest_contact_sq.c.member_id == models.LicenseHolder.id)
        .outerjoin(current_closure_sq, current_closure_sq.c.closure_id == models.LicenseHolder.closure_id)
    )

    if scope == "active":
        query = query.filter(_receivable_active_sql(current_closure_sq))
    elif scope == "closed":
        # 폐업/양도/이관 실제 Closure 기록이 있는 회원만 표시한다.
        query = query.filter(_receivable_closed_sql(current_closure_sq))

    # 명시적인 부과상태 필터가 있으면 그 필터를 우선한다.
    # 예: 미수금현황 탭(arrears_only=true)에서 '선납'을 선택했을 때
    # balance > 0 과 balance < 0 이 동시에 걸려 0건이 되는 충돌을 막는다.
    if arrears_only and not billing_status:
        query = query.filter(balance_expr > 0)
    if region:
        query = query.filter(models.LicenseHolder.region == region)
    if account_type:
        query = query.filter(ReceivableProfile.account_type == account_type)
    if contacted_only:
        query = query.filter(latest_contact_sq.c.member_id.isnot(None))
    if contact_status:
        if contact_status == "미연락":
            query = query.filter(or_(latest_contact_sq.c.status.is_(None), latest_contact_sq.c.status == "미연락"))
        else:
            query = query.filter(latest_contact_sq.c.status == contact_status)

    today_iso = datetime.now(KST).date().isoformat()
    pending_post_filter = billing_status == "pending"
    if billing_status == "pending":
        query = query.filter(
            _receivable_active_sql(current_closure_sq),
            ReceivableProfile.legacy_source_row.is_(None),
            ReceivableProfile.first_charge_date > today_iso,
        )
    elif billing_status == "arrears":
        query = query.filter(balance_expr > 0)
    elif billing_status == "settled":
        query = query.filter(balance_expr == 0)
    elif billing_status == "prepaid":
        query = query.filter(balance_expr < 0)

    raw_q = (q or "").strip()
    if raw_q:
        pat = f"%{raw_q}%"
        query = query.filter(
            or_(
                models.LicenseHolder.name.ilike(pat),
                models.LicenseHolder.vehicle_number.ilike(pat),
                models.LicenseHolder.management_number.ilike(pat),
                models.LicenseHolder.mobile.ilike(pat),
                current_closure_sq.c.management_number.ilike(pat),
                current_closure_sq.c.closure_type.ilike(pat),
                current_closure_sq.c.transferee.ilike(pat),
                current_closure_sq.c.transfer_region.ilike(pat),
            )
        )

    ordered = query.order_by(balance_expr.desc(), models.LicenseHolder.name.asc(), models.LicenseHolder.vehicle_number.asc())
    if pending_post_filter:
        # 후보 자체가 소수라서 실제 등록일을 Python에서 정확히 파싱 후 페이지네이션한다.
        candidate_rows = ordered.all()
        candidate_rows = [r for r in candidate_rows if _is_true_new_member(r[0], r[1])]
        total = len(candidate_rows)
        start = (page - 1) * limit
        rows = candidate_rows[start:start + limit]
    else:
        total = query.order_by(None).with_entities(func.count()).scalar() or 0
        rows = ordered.offset((page - 1) * limit).limit(limit).all()

    items = [
        _serialize_member(
            member,
            profile,
            int(balance or 0),
            contact_status or "미연락",
            last_contact_date or "",
            closure_id,
            closure_management_number or "",
            closure_date or "",
            closure_type or "",
            closure_reason or "",
            transferee or "",
            transfer_region or "",
            closure_receipt_date or "",
        )
        for (
            member, profile, balance, contact_status, last_contact_date,
            closure_id, closure_management_number, closure_date, closure_type,
            closure_reason, transferee, transfer_region, closure_receipt_date,
        ) in rows
    ]
    _schedule_background_sync(background_tasks)
    return {
        "items": items,
        "count": int(total),
        "page": page,
        "limit": limit,
        "pages": max(1, (int(total) + limit - 1) // limit),
    }



# ─────────────────────────────────────────────────────────────
# 통장 / 카드결제 일괄수납
# ─────────────────────────────────────────────────────────────

class ImportMatchIn(BaseModel):
    member_id: int


def _norm_col(v) -> str:
    return re.sub(r"[^0-9A-Za-z가-힣]", "", str(v or "")).lower()


def _digits(v) -> str:
    return re.sub(r"\D", "", str(v or ""))


def _amount_int(v) -> int:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return 0
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        try:
            return int(round(float(v)))
        except Exception:
            return 0
    text_value = str(v).strip().replace(",", "").replace("원", "").replace("₩", "")
    if not text_value:
        return 0
    neg = text_value.startswith("(") and text_value.endswith(")")
    text_value = re.sub(r"[^0-9.\-]", "", text_value)
    try:
        n = int(round(float(text_value)))
        return -abs(n) if neg else n
    except Exception:
        return 0


def _string_value(v) -> str:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return ""
    if isinstance(v, (datetime, date)):
        return v.isoformat()[:10]
    return str(v).strip()


_IMPORT_ALIASES = {
    "date": ["입금일", "거래일", "거래일자", "거래일시", "결제일", "결제일자", "승인일", "승인일시", "날짜", "일자"],
    "amount": ["입금액", "입금", "결제금액", "승인금액", "결제액", "거래금액", "금액"],
    "payer": ["입금자명", "입금자", "예금주", "성명", "구매자명", "고객명", "회원명", "거래내용", "적요", "내용"],
    "vehicle": ["차량번호", "차번", "차량", "자동차번호"],
    "management": ["관리번호", "회원관리번호"],
    "mobile": ["핸드폰", "핸드폰번호", "휴대폰", "휴대폰번호", "연락처", "전화번호"],
    "external": ["거래번호", "거래id", "승인번호", "결제번호", "주문번호", "거래고유번호", "transactionid"],
    "memo": ["메모", "비고", "적요", "거래내용", "내용", "상품명", "결제수단"],
}


def _non_receivable_import_reason(item) -> str:
    """자격증명/대폐차/가입비 등 월 미수금과 무관한 수입을 식별한다.

    자동으로 돈을 버리지 않고 `확인필요`로 분리한다. 복합입금(예: 자격증명+관리비)은
    전체 금액 자동반영을 금지하고 실제 회비 부분만 별도 수동입금하도록 한다.
    """
    if item is None:
        return ""
    if isinstance(item, str):
        text_blob = item
    else:
        raw = item.get("raw_data") or {} if isinstance(item, dict) else {}
        pieces = []
        if isinstance(item, dict):
            pieces.extend([item.get("memo", ""), item.get("payer_name", ""), item.get("external_id", "")])
        if isinstance(raw, dict):
            pieces.extend(str(v or "") for v in raw.values())
        text_blob = " ".join(str(x or "") for x in pieces)
    normalized = re.sub(r"\s+", "", str(text_blob or "")).lower()
    for keyword in NON_RECEIVABLE_IMPORT_KEYWORDS:
        if re.sub(r"\s+", "", keyword).lower() in normalized:
            return f"미수금 외 수입 의심({keyword})"
    return ""


def _find_import_header(raw: pd.DataFrame) -> int:
    best_idx, best_score = 0, -1
    max_rows = min(len(raw), 35)
    all_aliases = {_norm_col(a) for vals in _IMPORT_ALIASES.values() for a in vals}
    amount_aliases = {_norm_col(a) for a in _IMPORT_ALIASES["amount"]}
    date_aliases = {_norm_col(a) for a in _IMPORT_ALIASES["date"]}
    for idx in range(max_rows):
        vals = {_norm_col(x) for x in raw.iloc[idx].tolist() if _string_value(x)}
        score = sum(2 if v in amount_aliases or v in date_aliases else 1 for v in vals if v in all_aliases)
        if score > best_score:
            best_idx, best_score = idx, score
    return best_idx


def _read_import_frames(data: bytes, filename: str):
    ext = Path(filename or "").suffix.lower()
    frames = []
    if ext == ".csv":
        last_err = None
        for enc in ("utf-8-sig", "cp949", "euc-kr"):
            try:
                raw = pd.read_csv(io.BytesIO(data), header=None, dtype=object, encoding=enc)
                header = _find_import_header(raw)
                df = raw.iloc[header + 1:].copy()
                df.columns = [_string_value(x) or f"col_{i}" for i, x in enumerate(raw.iloc[header].tolist())]
                frames.append(("CSV", df))
                break
            except Exception as exc:
                last_err = exc
        if not frames and last_err:
            raise last_err
        return frames

    if ext not in {".xlsx", ".xls", ".xlsm"}:
        raise HTTPException(400, "지원 파일은 xlsx/xls/xlsm/csv 입니다.")
    try:
        excel = pd.ExcelFile(io.BytesIO(data))
        for sheet in excel.sheet_names:
            raw = pd.read_excel(excel, sheet_name=sheet, header=None, dtype=object)
            if raw.empty:
                continue
            header = _find_import_header(raw)
            df = raw.iloc[header + 1:].copy()
            df.columns = [_string_value(x) or f"col_{i}" for i, x in enumerate(raw.iloc[header].tolist())]
            frames.append((sheet, df))
    except ImportError as exc:
        raise HTTPException(400, "구형 .xls 파일 처리를 위해 xlrd 설치가 필요합니다. requirements.txt의 xlrd 항목을 함께 반영해주세요.") from exc
    except Exception as exc:
        raise HTTPException(400, f"파일을 읽지 못했습니다: {exc}") from exc
    return frames


def _pick_column(columns, aliases, prefer_exact=True):
    pairs = [(c, _norm_col(c)) for c in columns]
    for alias in aliases:
        na = _norm_col(alias)
        for c, nc in pairs:
            if nc == na:
                return c
    if not prefer_exact:
        for alias in aliases:
            na = _norm_col(alias)
            for c, nc in pairs:
                if na and na in nc:
                    return c
    return None


def _extract_vehicle_from_text(text_value: str) -> str:
    s = str(text_value or "")
    patterns = [
        r"(?:강원\s*)?\d{2,3}\s*[가-힣]\s*\d{4}\s*호?",
        r"\b\d{2,3}[- ]\d{4}\b",
    ]
    for pat in patterns:
        m = re.search(pat, s)
        if m:
            return m.group(0).strip()
    return ""


def _parse_import_rows(data: bytes, filename: str, source_type: str):
    parsed = []
    for sheet, df in _read_import_frames(data, filename):
        if df.empty:
            continue
        cols = list(df.columns)
        date_col = _pick_column(cols, _IMPORT_ALIASES["date"], prefer_exact=False)
        amount_col = _pick_column(cols, _IMPORT_ALIASES["amount"], prefer_exact=False)
        payer_col = _pick_column(cols, _IMPORT_ALIASES["payer"], prefer_exact=False)
        vehicle_col = _pick_column(cols, _IMPORT_ALIASES["vehicle"], prefer_exact=False)
        mgmt_col = _pick_column(cols, _IMPORT_ALIASES["management"], prefer_exact=False)
        mobile_col = _pick_column(cols, _IMPORT_ALIASES["mobile"], prefer_exact=False)
        ext_col = _pick_column(cols, _IMPORT_ALIASES["external"], prefer_exact=False)
        memo_col = _pick_column(cols, _IMPORT_ALIASES["memo"], prefer_exact=False)
        if amount_col is None:
            continue
        for local_idx, (_, row) in enumerate(df.iterrows(), start=1):
            amount = _amount_int(row.get(amount_col))
            if amount <= 0:
                continue
            raw_date = row.get(date_col) if date_col else None
            d = _parse_date(raw_date)
            if not d:
                try:
                    ts = pd.to_datetime(raw_date, errors="coerce")
                    d = None if pd.isna(ts) else ts.date()
                except Exception:
                    d = None
            payer = _string_value(row.get(payer_col)) if payer_col else ""
            memo = _string_value(row.get(memo_col)) if memo_col else ""
            vehicle = _string_value(row.get(vehicle_col)) if vehicle_col else ""
            if not vehicle:
                vehicle = _extract_vehicle_from_text(" ".join([payer, memo]))
            management = _string_value(row.get(mgmt_col)) if mgmt_col else ""
            mobile = _string_value(row.get(mobile_col)) if mobile_col else ""
            external = _string_value(row.get(ext_col)) if ext_col else ""
            raw_dict = {str(k): _string_value(v) for k, v in row.to_dict().items() if _string_value(v)}
            fingerprint_base = (
                f"{source_type}|id|{_norm_col(external)}" if external else
                "|".join([source_type, d.isoformat() if d else "", str(amount), _norm_col(payer), _norm_col(vehicle), _norm_col(memo)])
            )
            parsed.append({
                "sheet": sheet,
                "source_row": local_idx,
                "transaction_date": d.isoformat() if d else datetime.now(KST).date().isoformat(),
                "payer_name": payer,
                "amount": amount,
                "vehicle_number": vehicle,
                "management_number": management,
                "mobile": mobile,
                "external_id": external,
                "memo": memo,
                "fingerprint": hashlib.sha256(fingerprint_base.encode("utf-8")).hexdigest(),
                "raw_data": raw_dict,
            })
    if not parsed:
        raise HTTPException(400, "입금액/결제금액이 있는 거래를 찾지 못했습니다. 파일의 열 이름을 확인해주세요.")
    return parsed


def _member_match_maps(db: Session):
    members = (
        db.query(models.LicenseHolder)
        .filter(or_(models.LicenseHolder.status.is_(None), models.LicenseHolder.status != "pending"))
        .all()
    )

    def add(mp, key, member):
        if key:
            mp.setdefault(key, []).append(member)

    vehicles, names, mgmts, mobiles = {}, {}, {}, {}
    for m in members:
        add(vehicles, _norm(getattr(m, "vehicle_number", "")), m)
        add(names, _norm(getattr(m, "name", "")), m)
        add(mgmts, _norm(getattr(m, "management_number", "")), m)
        add(mobiles, _digits(getattr(m, "mobile", "")), m)

    aliases = {}
    hist = (
        db.query(ReceivableImportRow.payer_name, ReceivableImportRow.matched_member_id)
        .filter(ReceivableImportRow.status == "posted", ReceivableImportRow.matched_member_id.isnot(None))
        .all()
    )
    temp = {}
    for payer, mid in hist:
        key = _norm(payer)
        if key:
            temp.setdefault(key, set()).add(int(mid))
    for key, mids in temp.items():
        if len(mids) == 1:
            aliases[key] = next(iter(mids))
    by_id = {m.id: m for m in members}
    return by_id, vehicles, names, mgmts, mobiles, aliases


def _unique_member(candidates):
    if not candidates:
        return None
    uniq = {m.id: m for m in candidates}
    if len(uniq) == 1:
        return next(iter(uniq.values()))
    active = [m for m in uniq.values() if _is_active(m)]
    return active[0] if len(active) == 1 else None


def _auto_match_import(row, maps):
    by_id, vehicles, names, mgmts, mobiles, aliases = maps
    payer_key = _norm(row.get("payer_name"))

    # 과거 입금자 매칭이력보다 현재 거래에 들어 있는 관리번호/차량/핸드폰을 우선한다.
    # 동명이인(예: 이건우)이 생긴 뒤에도 과거 alias가 엉뚱한 사람에게 입금을 붙이지 않게 한다.
    mgmt_key = _norm(row.get("management_number"))
    m = _unique_member(mgmts.get(mgmt_key)) if mgmt_key else None
    if m:
        return m, "관리번호 정확일치"
    vehicle_key = _norm(row.get("vehicle_number"))
    m = _unique_member(vehicles.get(vehicle_key)) if vehicle_key else None
    if m:
        return m, "차량번호 정확일치"
    mobile_key = _digits(row.get("mobile"))
    m = _unique_member(mobiles.get(mobile_key)) if len(mobile_key) >= 8 else None
    if m:
        return m, "핸드폰 정확일치"

    # 입금자 과거 alias는 현재 동일 성명 회원이 정확히 1명일 때만 허용한다.
    # 같은 이름의 회원이 2명 이상이면 자동매칭하지 않고 확인필요로 남긴다.
    if payer_key and payer_key in aliases and aliases[payer_key] in by_id:
        name_candidates = {x.id for x in (names.get(payer_key) or [])}
        if len(name_candidates) == 1 and aliases[payer_key] in name_candidates:
            return by_id[aliases[payer_key]], "기존 입금자 매칭이력(동명이인 없음)"

    m = _unique_member(names.get(payer_key)) if payer_key else None
    if m:
        return m, "성명 유일일치"
    if payer_key and len(payer_key) >= 2:
        candidate_members = []
        for name_key, rows in names.items():
            if len(name_key) >= 2 and (payer_key.startswith(name_key) or payer_key.endswith(name_key)):
                candidate_members.extend(rows)
        m = _unique_member(candidate_members)
        if m:
            return m, "입금자명 포함 유일일치"
    return None, "확인 필요"


def _import_row_json(row: ReceivableImportRow, member_map=None):
    member = (member_map or {}).get(row.matched_member_id) if row.matched_member_id else None
    return {
        "id": row.id,
        "batch_id": row.batch_id,
        "source_row": row.source_row,
        "transaction_date": row.transaction_date or "",
        "payer_name": row.payer_name or "",
        "amount": int(row.amount or 0),
        "vehicle_number": row.vehicle_number or "",
        "management_number": row.management_number or "",
        "mobile": row.mobile or "",
        "external_id": row.external_id or "",
        "memo": row.memo or "",
        "matched_member_id": row.matched_member_id,
        "match_reason": row.match_reason or "",
        "status": row.status,
        "payment_id": row.payment_id,
        "matched_name": getattr(member, "name", "") if member else "",
        "matched_vehicle": getattr(member, "vehicle_number", "") if member else "",
        "matched_region": getattr(member, "region", "") if member else "",
    }


def _refresh_import_batch(db: Session, batch: ReceivableImportBatch):
    rows = db.query(ReceivableImportRow).filter(ReceivableImportRow.batch_id == batch.id).all()
    batch.total_rows = len(rows)
    batch.matched_rows = sum(1 for r in rows if r.status == "matched")
    batch.review_rows = sum(1 for r in rows if r.status == "review")
    batch.duplicate_rows = sum(1 for r in rows if r.status == "duplicate")
    batch.posted_rows = sum(1 for r in rows if r.status == "posted")
    # 반영예정액은 자동반영 확정(matched) + 이미 반영(posted)만 합산한다.
    # review(자격증명/발급비 의심 포함)는 금액 합계에도 넣지 않아 오해를 막는다.
    batch.total_amount = sum(int(r.amount or 0) for r in rows if r.status in {"matched", "posted"})
    batch.posted_amount = sum(int(r.amount or 0) for r in rows if r.status == "posted")
    return rows


@router.post("/api/receivables/imports/preview")
async def preview_payment_import(
    source_type: str = Form("통장"),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    _ensure_receivables_schema_ready()
    _ensure_db_ledger_ready(db)
    raw = await file.read()
    if not raw:
        raise HTTPException(400, "빈 파일입니다.")
    if len(raw) > 25 * 1024 * 1024:
        raise HTTPException(400, "파일은 25MB 이하만 업로드할 수 있습니다.")
    source_type = (source_type or "통장").strip()[:30]
    parsed = _parse_import_rows(raw, file.filename or "upload.xlsx", source_type)
    batch = ReceivableImportBatch(
        source_type=source_type,
        source_name=file.filename or "",
        status="preview",
        created_by=_user_name(current_user),
    )
    db.add(batch)
    db.flush()

    maps = _member_match_maps(db)
    fingerprints = [r["fingerprint"] for r in parsed]
    posted = set()
    for i in range(0, len(fingerprints), 800):
        chunk = fingerprints[i:i + 800]
        posted.update(
            fp for (fp,) in db.query(ReceivableImportRow.fingerprint)
            .filter(ReceivableImportRow.fingerprint.in_(chunk), ReceivableImportRow.status == "posted")
            .all()
        )

    seen = set()
    for item in parsed:
        duplicate = item["fingerprint"] in posted or item["fingerprint"] in seen
        seen.add(item["fingerprint"])
        non_receivable_reason = "" if duplicate else _non_receivable_import_reason(item)
        if duplicate:
            member, reason, status = None, "기존/파일내 중복", "duplicate"
        else:
            member, match_reason = _auto_match_import(item, maps)
            if non_receivable_reason:
                # 후보 회원은 보여주되, 자동반영 대상(matched)으로 만들지 않는다.
                reason = f"{non_receivable_reason} · 자동반영 금지" + (f" · 후보: {match_reason}" if member else "")
                status = "review"
            else:
                reason = match_reason
                status = "matched" if member else "review"
        db.add(ReceivableImportRow(
            batch_id=batch.id,
            source_row=item["source_row"],
            transaction_date=item["transaction_date"],
            payer_name=item["payer_name"],
            amount=item["amount"],
            vehicle_number=item["vehicle_number"],
            management_number=item["management_number"],
            mobile=item["mobile"],
            external_id=item["external_id"],
            memo=item["memo"],
            fingerprint=item["fingerprint"],
            matched_member_id=getattr(member, "id", None),
            match_reason=reason,
            status=status,
            raw_data={"sheet": item["sheet"], **item["raw_data"]},
        ))
    db.flush()
    _refresh_import_batch(db, batch)
    db.commit()
    return get_payment_import(batch.id, db=db, current_user=current_user)


@router.get("/api/receivables/imports/{batch_id}")
def get_payment_import(
    batch_id: int,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    _ensure_receivables_schema_ready()
    batch = db.query(ReceivableImportBatch).filter(ReceivableImportBatch.id == batch_id).first()
    if not batch:
        raise HTTPException(404, "업로드 내역을 찾을 수 없습니다.")
    rows = (
        db.query(ReceivableImportRow)
        .filter(ReceivableImportRow.batch_id == batch_id)
        .order_by(ReceivableImportRow.id)
        .all()
    )
    mids = {r.matched_member_id for r in rows if r.matched_member_id}
    member_map = {
        m.id: m for m in db.query(models.LicenseHolder).filter(models.LicenseHolder.id.in_(mids)).all()
    } if mids else {}
    return {
        "batch": {
            "id": batch.id,
            "source_type": batch.source_type,
            "source_name": batch.source_name or "",
            "status": batch.status,
            "total_rows": batch.total_rows,
            "matched_rows": batch.matched_rows,
            "review_rows": batch.review_rows,
            "duplicate_rows": batch.duplicate_rows,
            "posted_rows": batch.posted_rows,
            "total_amount": batch.total_amount,
            "posted_amount": batch.posted_amount,
        },
        "rows": [_import_row_json(r, member_map) for r in rows],
    }


@router.get("/api/receivables/import-member-search")
def import_member_search(
    q: str = Query(..., min_length=1),
    limit: int = Query(12, ge=1, le=30),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    term = f"%{q.strip()}%"
    rows = (
        db.query(models.LicenseHolder)
        .filter(or_(
            models.LicenseHolder.name.ilike(term),
            models.LicenseHolder.vehicle_number.ilike(term),
            models.LicenseHolder.management_number.ilike(term),
            models.LicenseHolder.mobile.ilike(term),
        ))
        .order_by(case((models.LicenseHolder.status == "active", 0), else_=1), models.LicenseHolder.name)
        .limit(limit)
        .all()
    )
    return {
        "items": [{
            "member_id": m.id,
            "name": m.name or "",
            "vehicle_number": m.vehicle_number or "",
            "region": m.region or "",
            "management_number": m.management_number or "",
            "active": _is_active(m),
        } for m in rows]
    }


@router.patch("/api/receivables/imports/rows/{row_id}/match")
def match_import_row(
    row_id: int,
    payload: ImportMatchIn,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    row = db.query(ReceivableImportRow).filter(ReceivableImportRow.id == row_id).first()
    if not row:
        raise HTTPException(404, "거래행을 찾을 수 없습니다.")
    member = db.query(models.LicenseHolder).filter(models.LicenseHolder.id == payload.member_id).first()
    if not member:
        raise HTTPException(404, "회원을 찾을 수 없습니다.")
    if row.status == "posted":
        raise HTTPException(400, "이미 수납 반영된 거래입니다.")
    row.matched_member_id = member.id
    reason = _non_receivable_import_reason({
        "payer_name": row.payer_name or "",
        "memo": row.memo or "",
        "external_id": row.external_id or "",
        "raw_data": row.raw_data or {},
    })
    if reason:
        # 자격증명+관리비 같은 복합입금은 전체 금액을 관리비로 올리면 안 된다.
        # 회원만 연결해 두고 확인필요 상태를 유지하며, 실제 회비 부분은 별도 수동입금한다.
        row.match_reason = f"수동연결 · {reason} · 금액분리 필요"
        if row.status != "duplicate":
            row.status = "review"
    else:
        row.match_reason = "수동연결"
        if row.status != "duplicate":
            row.status = "matched"
    batch = db.query(ReceivableImportBatch).filter(ReceivableImportBatch.id == row.batch_id).first()
    if batch:
        _refresh_import_batch(db, batch)
    db.commit()
    return {"ok": True}


@router.post("/api/receivables/imports/{batch_id}/post")
def post_payment_import(
    batch_id: int,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    _ensure_receivables_schema_ready()
    batch = db.query(ReceivableImportBatch).filter(ReceivableImportBatch.id == batch_id).first()
    if not batch:
        raise HTTPException(404, "업로드 내역을 찾을 수 없습니다.")
    rows = (
        db.query(ReceivableImportRow)
        .filter(ReceivableImportRow.batch_id == batch_id)
        .order_by(ReceivableImportRow.id)
        .all()
    )
    posted_fingerprints = {
        fp for (fp,) in db.query(ReceivableImportRow.fingerprint)
        .filter(ReceivableImportRow.status == "posted", ReceivableImportRow.batch_id != batch_id)
        .all()
    }
    posted_count = 0
    posted_amount = 0
    method = "카드" if "결제" in (batch.source_type or "") or "사이다" in (batch.source_type or "") else "계좌이체"
    for row in rows:
        # 안전장치: 일괄반영은 status == matched 인 거래만 허용한다.
        # review(자격증명/발급비 의심 포함)는 matched_member_id가 있더라도 절대 자동반영하지 않는다.
        if row.status != "matched":
            continue
        nonreceivable_reason = _non_receivable_import_reason({
            "payer_name": row.payer_name or "",
            "memo": row.memo or "",
            "external_id": row.external_id or "",
            "raw_data": row.raw_data or {},
        })
        if nonreceivable_reason:
            row.status = "review"
            row.match_reason = f"{nonreceivable_reason} · 자동반영 차단 · 금액분리 필요"
            continue
        if not row.matched_member_id:
            row.status = "review"
            continue
        if row.fingerprint in posted_fingerprints:
            row.status = "duplicate"
            row.match_reason = "기존 반영 거래와 중복"
            continue
        if not db.query(models.LicenseHolder.id).filter(models.LicenseHolder.id == row.matched_member_id).first():
            row.status = "review"
            row.match_reason = "회원 확인 필요"
            continue
        payment = ReceivablePayment(
            member_id=row.matched_member_id,
            payment_date=row.transaction_date or datetime.now(KST).date().isoformat(),
            amount=int(row.amount or 0),
            method=method,
            memo=(f"[일괄수납 #{batch.id}] {row.payer_name or ''} {row.memo or ''}").strip(),
            created_by=_user_name(current_user),
        )
        db.add(payment)
        db.flush()
        row.payment_id = payment.id
        row.status = "posted"
        row.match_reason = row.match_reason or "일괄반영"
        posted_fingerprints.add(row.fingerprint)
        posted_count += 1
        posted_amount += int(row.amount or 0)
    _refresh_import_batch(db, batch)
    batch.status = "posted" if batch.review_rows == 0 else "partial"
    batch.posted_at = datetime.now(KST)
    db.commit()
    return {
        "ok": True,
        "posted_rows": posted_count,
        "posted_amount": posted_amount,
        "review_rows": batch.review_rows,
        "duplicate_rows": batch.duplicate_rows,
    }



def _excel_safe_text(value) -> str:
    """Excel 수식 주입을 막으면서 화면에 보이는 문자열은 그대로 보존한다."""
    if value is None:
        return ""
    text_value = str(value)
    if text_value.startswith(("=", "+", "-", "@")):
        return "'" + text_value
    return text_value


def _export_view_label(view: str) -> str:
    return {
        "payment": "수납처리",
        "arrears": "미수금현황",
        "closed": "폐업관리",
        "contacts": "연락관리",
    }.get(view, "수납미수금")


@router.get("/api/receivables/export.xlsx")
def export_receivables_excel(
    view: str = Query("payment", pattern="^(payment|arrears|closed|contacts)$"),
    scope: str = Query("active", pattern="^(active|closed|all)$"),
    closure_mode: str = Query("current", pattern="^(current|history)$"),
    arrears_only: bool = False,
    q: str = "",
    region: str = "",
    account_type: str = "",
    contact_status: str = "",
    contacted_only: bool = False,
    billing_status: str = Query("", pattern="^(|pending|arrears|settled|prepaid)$"),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """현재 화면의 검색/지역/계정/부과상태/연락상태 필터를 그대로 적용해 전체 결과를 XLSX로 내려준다.

    화면의 50건 페이지가 아니라, 현재 필터에 해당하는 전체 건을 다운로드한다.
    회원마스터의 주소/핸드폰번호를 함께 포함하며 폐업현황은 Closure 스냅샷을 우선 사용한다.
    """
    _ensure_db_ledger_ready(db)

    # 기존 목록 API와 완전히 같은 판정/필터를 재사용한다. Python 내부 호출이므로
    # 외부 API의 limit<=200 제약과 무관하게 필터 결과 전체를 한 번에 받는다.
    result = list_members(
        background_tasks=BackgroundTasks(),
        scope=scope,
        closure_mode=closure_mode,
        arrears_only=arrears_only,
        q=q,
        region=region,
        account_type=account_type,
        contact_status=contact_status,
        contacted_only=contacted_only,
        billing_status=billing_status,
        page=1,
        limit=1000000,
        db=db,
        current_user=current_user,
    )
    items = result.get("items", [])

    member_ids = {int(x["member_id"]) for x in items if x.get("member_id") is not None}
    closure_ids = {int(x["closure_id"]) for x in items if x.get("closure_id") is not None}

    member_map = {}
    if member_ids:
        member_map = {
            m.id: m
            for m in db.query(models.LicenseHolder)
            .filter(models.LicenseHolder.id.in_(member_ids))
            .all()
        }

    closure_map = {}
    if closure_ids:
        closure_map = {
            c.id: c
            for c in db.query(models.Closure)
            .filter(models.Closure.id.in_(closure_ids))
            .all()
        }

    wb = Workbook()
    ws = wb.active
    sheet_label = _export_view_label(view)
    ws.title = sheet_label[:31]

    headers = [
        "성명", "계정", "차량번호", "지역", "주소", "핸드폰번호", "전화번호",
        "현재잔액", "미수금", "선납금", "부과상태", "현재상태",
        "연락상태", "최근연락일", "가입일자", "회원관리번호",
        "폐업관리번호", "구분", "접수일", "처리일", "양수인",
        "이관/양도지역", "사유",
    ]
    last_col = get_column_letter(len(headers))

    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = f"수납 · 미수금 - {sheet_label}"
    ws["A1"].font = Font(name="맑은 고딕", size=15, bold=True, color="1F2937")
    ws["A1"].fill = PatternFill("solid", fgColor="EEE9FF")
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 30

    billing_label = {
        "pending": "부과대기",
        "arrears": "미수",
        "settled": "완납",
        "prepaid": "선납",
    }.get(billing_status, "전체")
    filter_parts = [
        *(([f"폐업조회: {'현재 폐업자' if closure_mode == 'current' else '전체 폐업이력'}"] if view == "closed" else [])),
        f"검색: {(q or '').strip() or '전체'}",
        f"지역: {region or '전체'}",
        f"계정: {account_type or '전체'}",
        f"부과상태: {billing_label}",
        f"연락상태: {contact_status or '전체'}",
        f"결과: {len(items):,}{'건' if view == 'closed' else '명'}",
    ]
    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"] = " | ".join(filter_parts)
    ws["A2"].font = Font(name="맑은 고딕", size=10, color="667085")
    ws["A2"].fill = PatternFill("solid", fgColor="F8FAFC")

    ws.merge_cells(f"A3:{last_col}3")
    ws["A3"] = f"다운로드 기준: {datetime.now(KST).strftime('%Y-%m-%d %H:%M:%S')} · PostgreSQL 공식원장"
    ws["A3"].font = Font(name="맑은 고딕", size=9, color="98A2B3")

    header_row = 5
    for col_idx, label in enumerate(headers, 1):
        cell = ws.cell(header_row, col_idx, label)
        cell.font = Font(name="맑은 고딕", size=10, bold=True, color="344054")
        cell.fill = PatternFill("solid", fgColor="EAF2FF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[header_row].height = 24

    thin = Side(style="thin", color="E7EAF0")
    data_start = header_row + 1
    for row_idx, item in enumerate(items, data_start):
        mid = item.get("member_id")
        cid = item.get("closure_id")
        member = member_map.get(int(mid)) if mid is not None else None
        closure = closure_map.get(int(cid)) if cid is not None else None

        # 폐업관리 내보내기는 당시 폐업현황에 저장된 주소/연락처를 우선하고,
        # 비어 있으면 현재 회원마스터 값으로 보완한다.
        if view == "closed" and closure is not None:
            address = closure.address or (getattr(member, "address", "") if member else "")
            mobile = closure.mobile or (getattr(member, "mobile", "") if member else "")
            phone = closure.phone or (getattr(member, "phone", "") if member else "")
        else:
            address = getattr(member, "address", "") if member else (getattr(closure, "address", "") if closure else "")
            mobile = getattr(member, "mobile", "") if member else (getattr(closure, "mobile", "") if closure else "")
            phone = getattr(member, "phone", "") if member else (getattr(closure, "phone", "") if closure else "")

        balance = item.get("balance")
        if isinstance(balance, (int, float)):
            balance_num = int(balance)
            arrears_amount = max(balance_num, 0)
            prepaid_amount = max(-balance_num, 0)
        else:
            balance_num = None
            arrears_amount = None
            prepaid_amount = None

        if view == "closed":
            if closure_mode == "current":
                current_state = "폐업"
            else:
                current_state = "과거기록" if mid is None else ("현재활동" if item.get("active") else "폐업")
        else:
            current_state = "활성" if item.get("active") else "폐업"

        member_management_number = getattr(member, "management_number", "") if member else ""
        row_values = [
            _excel_safe_text(item.get("name", "")),
            _excel_safe_text(item.get("account_type", "")),
            _excel_safe_text(item.get("vehicle_number", "")),
            _excel_safe_text(item.get("region", "")),
            _excel_safe_text(address),
            _excel_safe_text(mobile),
            _excel_safe_text(phone),
            balance_num,
            arrears_amount,
            prepaid_amount,
            _excel_safe_text(item.get("billing_state", "")),
            current_state,
            _excel_safe_text(item.get("contact_status", "")),
            _excel_safe_text(item.get("last_contact_date", "")),
            _excel_safe_text(item.get("membership_date", "")),
            _excel_safe_text(member_management_number),
            _excel_safe_text(item.get("closure_management_number", "")),
            _excel_safe_text(item.get("closure_type", "")),
            _excel_safe_text(item.get("closure_receipt_date", "")),
            _excel_safe_text(item.get("closure_date", "")),
            _excel_safe_text(item.get("transferee", "")),
            _excel_safe_text(item.get("transfer_region", "")),
            _excel_safe_text(item.get("closure_reason", "")),
        ]
        for col_idx, value in enumerate(row_values, 1):
            cell = ws.cell(row_idx, col_idx, value)
            cell.font = Font(name="맑은 고딕", size=10, color="344054")
            cell.alignment = Alignment(vertical="center", wrap_text=(col_idx in (5, 23)))
            cell.border = Border(bottom=thin)

        # 금액은 숫자로 저장하여 사용자가 Excel에서 바로 합계/필터 가능하게 한다.
        for col_idx in (8, 9, 10):
            ws.cell(row_idx, col_idx).number_format = '#,##0;[Red]-#,##0'
            ws.cell(row_idx, col_idx).alignment = Alignment(horizontal="right", vertical="center")
        if balance_num is not None and balance_num < 0:
            ws.cell(row_idx, 8).font = Font(name="맑은 고딕", size=10, bold=True, color="2563EB")
            ws.cell(row_idx, 10).font = Font(name="맑은 고딕", size=10, bold=True, color="2563EB")
        elif balance_num is not None and balance_num > 0:
            ws.cell(row_idx, 8).font = Font(name="맑은 고딕", size=10, bold=True, color="C2413B")
            ws.cell(row_idx, 9).font = Font(name="맑은 고딕", size=10, bold=True, color="C2413B")

    data_end = max(header_row, header_row + len(items))
    ws.auto_filter.ref = f"A{header_row}:{last_col}{data_end}"
    ws.freeze_panes = f"A{data_start}"

    widths = {
        1: 14, 2: 11, 3: 18, 4: 11, 5: 38, 6: 16, 7: 15,
        8: 14, 9: 14, 10: 14, 11: 14, 12: 12, 13: 14, 14: 13,
        15: 13, 16: 15, 17: 15, 18: 10, 19: 13, 20: 13, 21: 15,
        22: 20, 23: 30,
    }
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.sheet_view.showGridLines = False
    ws.auto_filter.ref = f"A{header_row}:{last_col}{data_end}"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    stamp = datetime.now(KST).strftime("%Y%m%d_%H%M")
    filename = f"수납미수금_{sheet_label}_{stamp}.xlsx"
    encoded = quote(filename)
    response_headers = {
        "Content-Disposition": f"attachment; filename=receivables_{stamp}.xlsx; filename*=UTF-8''{encoded}",
        "Cache-Control": "no-store",
        "X-Export-Count": str(len(items)),
    }
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=response_headers,
    )


@router.get("/api/receivables/members/{member_id}")
def member_detail(
    member_id: int,
    year: int = 2026,
    closure_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    _ensure_db_ledger_ready(db)
    _ensure_current_month_billing(db)
    member = db.query(models.LicenseHolder).filter(models.LicenseHolder.id == member_id).first()
    if not member:
        raise HTTPException(404, "회원을 찾을 수 없습니다.")
    profile = _ensure_profile_for_member(db, member_id)
    if not profile:
        raise HTTPException(404, "미수금 프로필을 찾을 수 없습니다.")

    # 현재 폐업상태와 과거 폐업현황 조회 문맥을 분리한다.
    # active 회원에게 과거 closure 이력이 있어도 현재 폐업으로 취급하지 않는다.
    current_closure = None
    if (member.status or "active") == "closed":
        if getattr(member, "closure_id", None):
            current_closure = (
                db.query(models.Closure)
                .filter(
                    models.Closure.id == member.closure_id,
                    models.Closure.deleted_at.is_(None),
                )
                .first()
            )
        if current_closure is None:
            current_closure = (
                db.query(models.Closure)
                .filter(models.Closure.member_id == member_id, models.Closure.deleted_at.is_(None))
                .order_by(models.Closure.id.desc())
                .first()
            )

    selected_closure = None
    if closure_id is not None:
        candidate = (
            db.query(models.Closure)
            .filter(models.Closure.id == closure_id, models.Closure.deleted_at.is_(None))
            .first()
        )
        if candidate is not None:
            direct = getattr(candidate, "member_id", None) == member_id
            same_identity = (
                _norm(candidate.name) == _norm(member.name)
                and _norm(candidate.vehicle_number) == _norm(member.vehicle_number)
                and bool(_norm(candidate.name)) and bool(_norm(candidate.vehicle_number))
            )
            if direct or same_identity:
                selected_closure = candidate

    # 계산/자동부과 중단에는 현재 폐업건만 사용한다.
    closure = current_closure

    # 과거 버그로 DB에 남아 있어도 미래/legacy중복/폐업후 auto charge는 상세 잔액에서 제외한다.
    all_member_charges = db.query(ReceivableCharge).filter(ReceivableCharge.member_id == member_id).all()
    valid_member_charges = [
        ch for ch in all_member_charges
        if ch.source != "auto" or _valid_auto_charge(profile, member, closure, ch)
    ]
    charge_total = sum(int(ch.amount or 0) for ch in valid_member_charges)

    payment_total_query = db.query(func.coalesce(func.sum(ReceivablePayment.amount), 0)).filter(
        ReceivablePayment.member_id == member_id,
        ReceivablePayment.cancelled_at.is_(None),
    )
    if profile.legacy_source_row is not None:
        cutoff = _legacy_snapshot_cutoff_at()
        payment_total_query = payment_total_query.filter(or_(
            ReceivablePayment.payment_date > LEGACY_DATA_THROUGH_DATE_ISO,
            ReceivablePayment.created_at > cutoff,
        ))
    payment_total = payment_total_query.scalar() or 0
    baseline_balance = _legacy_balance_as_of(profile, _parse_date(getattr(closure, "closure_date", None)) if closure else None)
    balance = int(baseline_balance) + int(charge_total) - int(payment_total)

    latest = (
        db.query(ReceivableContactLog)
        .filter(ReceivableContactLog.member_id == member_id)
        .order_by(ReceivableContactLog.contact_date.desc(), ReceivableContactLog.id.desc())
        .first()
    )

    program_charges = [
        ch for ch in valid_member_charges
        if str(ch.billing_month or "").startswith(f"{year}-")
    ]
    program_payments_query = (
        db.query(ReceivablePayment)
        .filter(
            ReceivablePayment.member_id == member_id,
            ReceivablePayment.cancelled_at.is_(None),
            ReceivablePayment.payment_date.like(f"{year}-%"),
        )
    )
    if profile.legacy_source_row is not None:
        cutoff = _legacy_snapshot_cutoff_at()
        program_payments_query = program_payments_query.filter(or_(
            ReceivablePayment.payment_date > LEGACY_DATA_THROUGH_DATE_ISO,
            ReceivablePayment.created_at > cutoff,
        ))
    program_payments = program_payments_query.all()
    ch_by_month = {int(c.billing_month[-2:]): int(c.amount) for c in program_charges}
    pay_by_month = {}
    pay_dates = {}
    adjustment_by_month = {}
    adjustment_dates = {}
    for p in program_payments:
        try:
            mm = int(p.payment_date[5:7])
        except Exception:
            continue
        if (p.method or "").strip() == "잔액수정":
            # ReceivablePayment는 잔액에서 빼는 구조이므로, 저장된 signed amount의 반대값이 실제 잔액 조정효과다.
            effect = -int(p.amount or 0)
            adjustment_by_month[mm] = adjustment_by_month.get(mm, 0) + effect
            adjustment_dates.setdefault(mm, []).append(p.payment_date)
        else:
            pay_by_month[mm] = pay_by_month.get(mm, 0) + int(p.amount)
            pay_dates.setdefault(mm, []).append(p.payment_date)

    legacy_by_month = {}
    legacy_seed = _seed_for_profile(profile) if year == 2026 else None
    legacy_carryover = int((legacy_seed or {}).get("carryover") or 0) if year == 2026 else 0
    if year == 2026:
        for row in profile.legacy_months or []:
            legacy_by_month[int(row.get("month") or 0)] = row

    if year <= 2026:
        running = 0
    else:
        before_valid_charges = sum(
            int(ch.amount or 0)
            for ch in valid_member_charges
            if str(ch.billing_month or "") < f"{year}-01"
        )
        before_payments_query = (
            db.query(func.coalesce(func.sum(ReceivablePayment.amount), 0))
            .filter(
                ReceivablePayment.member_id == member_id,
                ReceivablePayment.cancelled_at.is_(None),
                ReceivablePayment.payment_date < f"{year}-01-01",
            )
        )
        if profile.legacy_source_row is not None:
            cutoff = _legacy_snapshot_cutoff_at()
            before_payments_query = before_payments_query.filter(or_(
                ReceivablePayment.payment_date > LEGACY_DATA_THROUGH_DATE_ISO,
                ReceivablePayment.created_at > cutoff,
            ))
        before_payments = before_payments_query.scalar() or 0
        running = int(profile.legacy_balance or 0) + int(before_valid_charges) - int(before_payments)

    monthly = []
    today = datetime.now(KST).date()
    current_month_key = _month_key(today)
    close_d = _parse_date(getattr(closure, "closure_date", None)) if closure else None
    close_month_key = _month_key(close_d) if close_d else None
    previous_legacy_arrears = legacy_carryover

    for m in range(1, 13):
        legacy = legacy_by_month.get(m) or {}
        month_key = f"{year}-{m:02d}"
        has_legacy_row = (
            any(legacy.get(k) is not None for k in ("billed_total", "payment", "arrears"))
            or bool(legacy.get("payment_date"))
            or bool(legacy.get("monthly_charge"))
        )

        # Excel의 '월 부과금' 열은 실제 월회비가 아니라 전월 잔액이 포함된 누적 청구액이다.
        # 화면에는 사용자가 이해하는 실제 월 부과액만 표시한다.
        legacy_monthly_charge = legacy.get("monthly_charge")
        if legacy_monthly_charge is None and legacy.get("billed_total") is not None:
            # 구버전 snapshot 호환. 월부과액은 누적청구액 차이가 아니라 계정별 고정요율로 표시한다.
            expected_fee = ACCOUNT_FEES.get(profile.account_type)
            legacy_monthly_charge = int(expected_fee) if expected_fee is not None else None

        # 원본에 월말 미수금이 비어 있는 달은 청구금/고정월부과액으로 복원한다.
        # 예: 4~7월 미수칸이 공란이어도 월 10,000원씩 정상 증가하도록 표시.
        reconstructed_legacy = None
        if year == LEGACY_YEAR and m <= LEGACY_DATA_THROUGH_MONTH:
            reconstructed_legacy = _legacy_reconstructed_balance(profile, m)
            if reconstructed_legacy is not None:
                running = int(reconstructed_legacy)

        auto_charge = int(ch_by_month.get(m, 0) or 0)
        extra_paid = int(pay_by_month.get(m, 0) or 0)
        balance_adjustment = int(adjustment_by_month.get(m, 0) or 0)
        running += auto_charge
        running -= extra_paid
        running += balance_adjustment

        # 미래월/폐업 이후 월에 아무 원본·프로그램 활동이 없으면 '현재 미수금'을 만들어내지 않는다.
        inactive_future = month_key > current_month_key
        after_closure = bool(close_month_key and month_key > close_month_key)
        no_program_activity = auto_charge == 0 and extra_paid == 0 and balance_adjustment == 0
        display_current = None if (not has_legacy_row and no_program_activity and (inactive_future or after_closure)) else running

        monthly.append(
            {
                "month": m,
                "legacy_billed_total": legacy.get("billed_total"),
                "legacy_monthly_charge": legacy_monthly_charge,
                "legacy_payment": legacy.get("payment"),
                "legacy_payment_date": legacy.get("payment_date") or "",
                "legacy_arrears": (
                    legacy.get("arrears")
                    if legacy.get("arrears") is not None
                    else (reconstructed_legacy if has_legacy_row else None)
                ),
                "legacy_adjustment": legacy.get("legacy_adjustment"),
                "auto_charge": auto_charge,
                "additional_payment": extra_paid,
                "additional_payment_dates": pay_dates.get(m, []),
                "balance_adjustment": balance_adjustment,
                "balance_adjustment_dates": adjustment_dates.get(m, []),
                "current_arrears": display_current,
            }
        )

    # 상세의 "추가 입금·금액수정" 이력에도 snapshot에 이미 흡수된 과거 DB 수납을
    # 다시 노출하지 않는다. 원본 legacy 월별 입금은 위 월별 장부의 legacy_payment로 표시된다.
    payments_query = db.query(ReceivablePayment).filter(
        ReceivablePayment.member_id == member_id,
        ReceivablePayment.cancelled_at.is_(None),
    )
    if profile.legacy_source_row is not None:
        cutoff = _legacy_snapshot_cutoff_at()
        payments_query = payments_query.filter(or_(
            ReceivablePayment.payment_date > LEGACY_DATA_THROUGH_DATE_ISO,
            ReceivablePayment.created_at > cutoff,
        ))
    payments = (
        payments_query
        .order_by(ReceivablePayment.payment_date.desc(), ReceivablePayment.id.desc())
        .limit(100)
        .all()
    )
    contacts = (
        db.query(ReceivableContactLog)
        .filter(ReceivableContactLog.member_id == member_id)
        .order_by(ReceivableContactLog.contact_date.desc(), ReceivableContactLog.id.desc())
        .limit(100)
        .all()
    )

    member_json = _serialize_member(
        member,
        profile,
        balance,
        latest.status if latest else "미연락",
        latest.contact_date if latest else "",
        getattr(current_closure, "id", None),
        getattr(current_closure, "management_number", "") or "",
        getattr(current_closure, "closure_date", "") or "",
        getattr(current_closure, "closure_type", "") or "",
        getattr(current_closure, "reason", "") or "",
        getattr(current_closure, "transferee", "") or "",
        getattr(current_closure, "transfer_region", "") or "",
        getattr(current_closure, "receipt_date", "") or "",
    )
    if selected_closure is not None:
        member_json["closure_context"] = {
            "id": selected_closure.id,
            "management_number": selected_closure.management_number or "",
            "closure_type": _normalize_closure_type(selected_closure.closure_type),
            "closure_date": selected_closure.closure_date or "",
            "receipt_date": selected_closure.receipt_date or "",
            "transferee": selected_closure.transferee or "",
            "transfer_region": selected_closure.transfer_region or "",
            "reason": selected_closure.reason or "",
        }
    else:
        member_json["closure_context"] = None
    return {
        "member": member_json,
        "profile": {
            "account_type": profile.account_type,
            "unit_fee": int(profile.unit_fee or 0),
            "vehicle_count": int(profile.vehicle_count or 1),
            "first_charge_date": profile.first_charge_date or "",
            "legacy_balance": int(profile.legacy_balance or 0),
            "legacy_carryover": int(legacy_carryover or 0),
            "legacy_note": profile.legacy_note or "",
        },
        "monthly": monthly,
        "payments": [
            {
                "id": p.id,
                "payment_date": p.payment_date,
                "amount": p.amount,
                "method": p.method or "",
                "memo": p.memo or "",
                "created_by": p.created_by or "",
                "is_balance_edit": (p.method or "").strip() == "잔액수정",
                "balance_effect": -int(p.amount or 0) if (p.method or "").strip() == "잔액수정" else None,
            }
            for p in payments
        ],
        "contacts": [
            {
                "id": c.id,
                "contact_date": c.contact_date,
                "contact_method": c.contact_method,
                "status": c.status,
                "memo": c.memo or "",
                "created_by": c.created_by or "",
            }
            for c in contacts
        ],
    }


@router.post("/api/receivables/members/{member_id}/payments")
def add_payment(
    member_id: int,
    payload: PaymentIn,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    _ensure_db_ledger_ready(db)
    d = _parse_date(payload.payment_date)
    if not d:
        raise HTTPException(400, "입금일 형식이 올바르지 않습니다.")
    profile = _ensure_profile_for_member(db, member_id)
    if not profile:
        raise HTTPException(404, "미수금 프로필을 찾을 수 없습니다.")
    non_receivable_reason = _non_receivable_import_reason({"memo": payload.memo or ""})
    if non_receivable_reason:
        raise HTTPException(400, f"{non_receivable_reason}은(는) 관리비/협회비 미수금 수납으로 등록할 수 없습니다.")
    row = ReceivablePayment(
        member_id=member_id,
        payment_date=d.isoformat(),
        amount=int(payload.amount),
        method=(payload.method or "").strip() or None,
        memo=(payload.memo or "").strip() or None,
        created_by=_user_name(current_user),
    )
    db.add(row)
    db.commit()
    db.refresh(row)
    return {"ok": True, "payment_id": row.id}


@router.delete("/api/receivables/payments/{payment_id}")
def cancel_payment(
    payment_id: int,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    row = db.query(ReceivablePayment).filter(ReceivablePayment.id == payment_id).first()
    if not row:
        raise HTTPException(404, "입금내역을 찾을 수 없습니다.")
    if row.cancelled_at is None:
        row.cancelled_at = datetime.now(KST)
        row.cancelled_by = _user_name(current_user)
        db.commit()
    return {"ok": True}


@router.patch("/api/receivables/members/{member_id}/balance")
def edit_current_balance(
    member_id: int,
    payload: BalanceEditIn,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """현재 미수/선납 금액을 감사이력이 남는 방식으로 정정한다.

    실제 입금과 섞이지 않도록 별도 `잔액수정` 레코드로 남긴다.
    기존 잔액 산식(기준원장 + 부과 - 수납)을 그대로 이용하기 위해
    ReceivablePayment에 signed correction을 기록하되, 오늘 수납 KPI에서는 제외한다.
    """
    _ensure_db_ledger_ready(db)
    _ensure_current_month_billing(db)
    member = db.query(models.LicenseHolder).filter(models.LicenseHolder.id == member_id).first()
    if not member:
        raise HTTPException(404, "회원을 찾을 수 없습니다.")
    profile = _ensure_profile_for_member(db, member_id)
    if not profile:
        raise HTTPException(404, "미수금 프로필을 찾을 수 없습니다.")

    kind = (payload.balance_type or "").strip()
    if kind not in {"미수금", "완납", "선납"}:
        raise HTTPException(400, "금액 구분은 미수금/완납/선납 중 하나여야 합니다.")
    if kind == "완납":
        target_balance = 0
    elif kind == "선납":
        target_balance = -int(payload.amount or 0)
    else:
        target_balance = int(payload.amount or 0)

    d = _parse_date(payload.effective_date) if payload.effective_date else datetime.now(KST).date()
    if not d:
        raise HTTPException(400, "수정 기준일 형식이 올바르지 않습니다.")
    reason = (payload.reason or "").strip()
    if len(reason) < 2:
        raise HTTPException(400, "수정 사유를 입력해주세요.")

    current_closure = None
    if (member.status or "active") == "closed":
        if getattr(member, "closure_id", None):
            current_closure = db.query(models.Closure).filter(
                models.Closure.id == member.closure_id, models.Closure.deleted_at.is_(None)
            ).first()
        if current_closure is None:
            current_closure = db.query(models.Closure).filter(
                models.Closure.member_id == member_id, models.Closure.deleted_at.is_(None)
            ).order_by(models.Closure.id.desc()).first()

    all_member_charges = db.query(ReceivableCharge).filter(ReceivableCharge.member_id == member_id).all()
    valid_member_charges = [
        ch for ch in all_member_charges
        if ch.source != "auto" or _valid_auto_charge(profile, member, current_closure, ch)
    ]
    charge_total = sum(int(ch.amount or 0) for ch in valid_member_charges)
    payment_total_query = db.query(func.coalesce(func.sum(ReceivablePayment.amount), 0)).filter(
        ReceivablePayment.member_id == member_id, ReceivablePayment.cancelled_at.is_(None)
    )
    if profile.legacy_source_row is not None:
        cutoff = _legacy_snapshot_cutoff_at()
        payment_total_query = payment_total_query.filter(or_(
            ReceivablePayment.payment_date > LEGACY_DATA_THROUGH_DATE_ISO,
            ReceivablePayment.created_at > cutoff,
        ))
    payment_total = payment_total_query.scalar() or 0
    baseline_balance = _legacy_balance_as_of(
        profile, _parse_date(getattr(current_closure, "closure_date", None)) if current_closure else None
    )
    current_balance = int(baseline_balance) + int(charge_total) - int(payment_total)
    if current_balance == target_balance:
        return {"ok": True, "changed": False, "old_balance": current_balance, "new_balance": target_balance}

    # balance = baseline + charges - payments
    # 따라서 목표잔액으로 이동하려면 payment signed amount = 현재잔액 - 목표잔액.
    signed_payment_amount = int(current_balance) - int(target_balance)
    row = ReceivablePayment(
        member_id=member_id,
        payment_date=d.isoformat(),
        amount=signed_payment_amount,
        method="잔액수정",
        memo=f"[금액수정] {current_balance:,}원 → {target_balance:,}원 | 사유: {reason}",
        created_by=_user_name(current_user),
    )
    db.add(row)
    db.commit()
    db.refresh(row)
    return {
        "ok": True,
        "changed": True,
        "adjustment_id": row.id,
        "old_balance": current_balance,
        "new_balance": target_balance,
        "balance_effect": target_balance - current_balance,
    }


@router.post("/api/receivables/members/{member_id}/contacts")
def add_contact(
    member_id: int,
    payload: ContactIn,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    _ensure_db_ledger_ready(db)
    if payload.status not in CONTACT_STATUSES:
        raise HTTPException(400, "지원하지 않는 연락상태입니다.")
    d = _parse_date(payload.contact_date)
    if not d:
        raise HTTPException(400, "연락일 형식이 올바르지 않습니다.")
    if not db.query(models.LicenseHolder).filter(models.LicenseHolder.id == member_id).first():
        raise HTTPException(404, "회원을 찾을 수 없습니다.")
    row = ReceivableContactLog(
        member_id=member_id,
        contact_date=d.isoformat(),
        contact_method=payload.contact_method.strip() or "전화",
        status=payload.status,
        memo=(payload.memo or "").strip() or None,
        created_by=_user_name(current_user),
    )
    db.add(row)
    db.commit()
    db.refresh(row)
    return {"ok": True, "contact_id": row.id}


@router.patch("/api/receivables/members/{member_id}/account")
def update_account(
    member_id: int,
    payload: AccountIn,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    _ensure_db_ledger_ready(db)
    if payload.account_type not in ACCOUNT_FEES:
        raise HTTPException(400, "계정은 협회비/관리비/70세 중 하나여야 합니다.")
    profile = _ensure_profile_for_member(db, member_id)
    if not profile:
        raise HTTPException(404, "미수금 프로필을 찾을 수 없습니다.")
    profile.account_type = payload.account_type
    profile.unit_fee = ACCOUNT_FEES[payload.account_type]
    profile.vehicle_count = 1
    profile.account_manual_override = 1
    db.commit()
    return {"ok": True}
